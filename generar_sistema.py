"""
Generador del Sistema Administrativo - Mudanzas Metepec
Ejecutar: python3 generar_sistema.py
Genera: MudanzasMetepec_Sistema.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
import openpyxl

# ─── PALETA DE COLORES ────────────────────────────────────────────────────────
AZUL_OSC   = "1A3C5E"
AZUL_MED   = "2E6DA4"
AZUL_CLAR  = "D6E4F0"
VERDE_OSC  = "1E6B3C"
VERDE_MED  = "27AE60"
VERDE_CLAR = "D5F5E3"
ROJO_OSC   = "922B21"
ROJO_CLAR  = "FADBD8"
AMBAR      = "F39C12"
AMBAR_CLAR = "FEF9E7"
GRIS_OSC   = "2C3E50"
GRIS_MED   = "7F8C8D"
GRIS_CLAR  = "F2F3F4"
BLANCO     = "FFFFFF"
NEGRO      = "000000"

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=NEGRO, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=AZUL_MED)
    return Border(left=s, right=s, top=s, bottom=s)

def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_h(ws, row, height):
    ws.row_dimensions[row].height = height

def header_cell(ws, row, col, text, bg=AZUL_OSC, fg=BLANCO, size=10,
                bold=True, h_align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=fg)
    c.alignment = align(h=h_align, v="center", wrap=True)
    c.border = border_thin()
    return c

def data_cell(ws, row, col, value=None, bg=BLANCO, bold=False,
              h_align="left", num_format=None, formula=None):
    c = ws.cell(row=row, column=col, value=value if formula is None else formula)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=GRIS_OSC)
    c.alignment = align(h=h_align, v="center")
    c.border = border_thin()
    if num_format:
        c.number_format = num_format
    return c

def merge_header(ws, row, col_start, col_end, text, bg=AZUL_OSC, fg=BLANCO,
                 size=12):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, size=size, color=fg)
    c.alignment = align(h="center", v="center")
    c.border = border_medium()
    return c

def money(ws, row, col, formula=None, value=None, bg=BLANCO):
    v = formula if formula else value
    c = ws.cell(row=row, column=col, value=v)
    c.fill = fill(bg)
    c.font = font(color=GRIS_OSC)
    c.alignment = align(h="right", v="center")
    c.border = border_thin()
    c.number_format = '"$"#,##0.00'
    return c

def pct(ws, row, col, formula, bg=BLANCO):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = fill(bg)
    c.font = font(color=GRIS_OSC)
    c.alignment = align(h="right", v="center")
    c.border = border_thin()
    c.number_format = '0.0"%"'
    return c

def add_dropdown(ws, formula, sqref):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = sqref
    ws.add_data_validation(dv)

# ─── HOJA 1: AGENDA DE SERVICIOS ──────────────────────────────────────────────

def build_agenda(wb):
    ws = wb.create_sheet("Agenda")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Título
    merge_header(ws, 1, 1, 22, "MUDANZAS METEPEC — AGENDA DE SERVICIOS",
                 bg=AZUL_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 22,
                 "Registro de todos los servicios: WhatsApp → Cobro → Análisis",
                 bg=AZUL_MED, size=10)
    row_h(ws, 2, 22)

    # Encabezados
    headers = [
        ("FOLIO", 8),
        ("F.REGISTRO", 12),
        ("F.SERVICIO", 12),
        ("HORA", 8),
        ("CLIENTE", 20),
        ("TELÉFONO", 13),
        ("TIPO", 16),
        ("ORIGEN", 20),
        ("DESTINO", 20),
        ("KM BASE", 9),
        ("UNIDAD", 16),
        ("PERSONAL", 22),
        ("DESCRIPCIÓN", 28),
        ("COTIZADO", 11),
        ("ANTICIPO", 11),
        ("SALDO", 11),
        ("ESTADO", 14),
        ("PAGO", 13),
        ("FACTURA", 9),
        ("CANT.PERSONAS", 10),
        ("FOLIO COSTO", 10),
        ("OBSERVACIONES", 30),
    ]
    for i, (h, w) in enumerate(headers, 1):
        header_cell(ws, 3, i, h)
        col_w(ws, i, w)
    row_h(ws, 3, 30)

    # 200 filas de datos con fórmulas
    for r in range(4, 204):
        row_bg = BLANCO if r % 2 == 0 else GRIS_CLAR

        # Folio automático
        c = ws.cell(row=r, column=1,
                    value=f'=IF(E{r}<>"","MUD-"&TEXT(ROW()-3,"000"),"")')
        c.fill = fill(AZUL_CLAR)
        c.font = font(bold=True, color=AZUL_OSC)
        c.alignment = align(h="center")
        c.border = border_thin()

        # Fecha registro (auto si hay cliente)
        c2 = ws.cell(row=r, column=2,
                     value=f'=IF(E{r}<>"",TODAY(),"")')
        c2.fill = fill(row_bg)
        c2.font = font(color=GRIS_OSC)
        c2.alignment = align(h="center")
        c2.border = border_thin()
        c2.number_format = "DD/MM/YYYY"

        for col in [3]:  # Fecha servicio
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.number_format = "DD/MM/YYYY"
            c.alignment = align(h="center")

        for col in [4]:  # Hora
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.number_format = "HH:MM"
            c.alignment = align(h="center")

        for col in [5, 6, 8, 9, 12, 13, 22]:  # texto
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="left", wrap=(col in [12, 13, 22]))

        for col in [10, 20]:  # numéricos
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="center")
            c.number_format = "#,##0.0"

        for col in [14, 15]:  # dinero
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.number_format = '"$"#,##0.00'
            c.alignment = align(h="right")

        # Saldo = cotizado - anticipo
        c_saldo = ws.cell(row=r, column=16,
                          value=f'=IF(N{r}<>"",N{r}-O{r},"")')
        c_saldo.fill = fill(AMBAR_CLAR)
        c_saldo.font = font(bold=True, color=GRIS_OSC)
        c_saldo.border = border_thin()
        c_saldo.number_format = '"$"#,##0.00'
        c_saldo.alignment = align(h="right")

        for col in [7, 11, 17, 18, 19]:  # dropdowns
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="center")

        # Folio costo (ref)
        c = ws.cell(row=r, column=21,
                    value=f'=IF(E{r}<>"","MUD-"&TEXT(ROW()-3,"000"),"")')
        c.fill = fill(AZUL_CLAR)
        c.font = font(color=AZUL_OSC)
        c.border = border_thin()
        c.alignment = align(h="center")

        row_h(ws, r, 18)

    # Dropdowns
    add_dropdown(ws,
        '"Mudanza local,Flete,Renta de unidad,Foráneo,Maniobra especial"',
        f"G4:G203")
    add_dropdown(ws,
        '"Pick Up,1.5 Cerrada,2.5,3.5,Varias unidades"',
        f"K4:K203")
    add_dropdown(ws,
        '"Pendiente,Confirmado,En proceso,Terminado,Cancelado"',
        f"Q4:Q203")
    add_dropdown(ws,
        '"Efectivo,Transferencia,Tarjeta,Mixto"',
        f"R4:R203")
    add_dropdown(ws, '"Sí,No"', f"S4:S203")

    # Formato condicional de estado
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill("solid", fgColor=VERDE_CLAR)
    red_fill   = PatternFill("solid", fgColor=ROJO_CLAR)
    amber_fill = PatternFill("solid", fgColor=AMBAR_CLAR)
    blue_fill  = PatternFill("solid", fgColor=AZUL_CLAR)

    ws.conditional_formatting.add(f"Q4:Q203",
        CellIsRule(operator="equal", formula=['"Terminado"'],
                   fill=green_fill, font=Font(color=VERDE_OSC, bold=True)))
    ws.conditional_formatting.add(f"Q4:Q203",
        CellIsRule(operator="equal", formula=['"Cancelado"'],
                   fill=red_fill, font=Font(color=ROJO_OSC, bold=True)))
    ws.conditional_formatting.add(f"Q4:Q203",
        CellIsRule(operator="equal", formula=['"En proceso"'],
                   fill=amber_fill, font=Font(color=AMBAR, bold=True)))
    ws.conditional_formatting.add(f"Q4:Q203",
        CellIsRule(operator="equal", formula=['"Confirmado"'],
                   fill=blue_fill, font=Font(color=AZUL_MED, bold=True)))

    return ws


# ─── HOJA 2: COSTOS DEL SERVICIO ──────────────────────────────────────────────

def build_costos(wb):
    ws = wb.create_sheet("Costos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    merge_header(ws, 1, 1, 20, "MUDANZAS METEPEC — COSTOS DEL SERVICIO",
                 bg=VERDE_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 20,
                 "Conectado por Folio con Agenda  |  Costo $7/km Pick Up · $9/km 1.5C · $10/km 2.5 · $12/km 3.5",
                 bg=VERDE_MED, size=10)
    row_h(ws, 2, 22)

    headers = [
        ("FOLIO", 10),
        ("CLIENTE", 20),
        ("UNIDAD", 14),
        ("KM REALES", 10),
        ("$/KM", 8),
        ("COSTO UNIDAD", 13),
        ("CHOFER", 12),
        ("AYUDANTES", 10),
        ("MANIOBRAS", 10),
        ("CASETAS", 10),
        ("COMIDAS", 10),
        ("HOTELES", 10),
        ("GASOLINA+", 10),
        ("MATERIALES", 11),
        ("OTROS", 10),
        ("TOTAL GASTOS", 13),
        ("PRECIO COBRADO", 14),
        ("UTILIDAD", 12),
        ("MARGEN %", 10),
        ("NOTAS", 22),
    ]
    for i, (h, w) in enumerate(headers, 1):
        header_cell(ws, 3, i, h, bg=VERDE_OSC)
        col_w(ws, i, w)
    row_h(ws, 3, 30)

    # Tabla de costos por km (fuera del área de datos, columna 22+)
    ws.cell(row=5, column=22, value="Tabla $/km")
    ws.cell(row=5, column=22).font = font(bold=True, color=VERDE_OSC)
    km_rates = [("Pick Up", 7), ("1.5 Cerrada", 9), ("2.5", 10), ("3.5", 12)]
    for i, (u, r_) in enumerate(km_rates, 6):
        ws.cell(row=i, column=22, value=u)
        ws.cell(row=i, column=23, value=r_)
        ws.cell(row=i, column=23).number_format = '"$"#,##0'

    for r in range(4, 204):
        row_bg = BLANCO if r % 2 == 0 else GRIS_CLAR

        # Folio (manual - pegar desde agenda)
        c = ws.cell(row=r, column=1)
        c.fill = fill(AZUL_CLAR)
        c.font = font(bold=True, color=AZUL_OSC)
        c.alignment = align(h="center")
        c.border = border_thin()

        # Cliente (VLOOKUP desde Agenda)
        c2 = ws.cell(row=r, column=2,
                     value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:E,5,0),"—"),"")')
        c2.fill = fill(row_bg)
        c2.border = border_thin()
        c2.alignment = align(h="left")

        # Unidad (VLOOKUP desde Agenda)
        c3 = ws.cell(row=r, column=3,
                     value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:K,11,0),"—"),"")')
        c3.fill = fill(row_bg)
        c3.border = border_thin()
        c3.alignment = align(h="center")

        # KM reales
        c4 = ws.cell(row=r, column=4)
        c4.fill = fill(row_bg)
        c4.border = border_thin()
        c4.alignment = align(h="right")
        c4.number_format = "#,##0.0"

        # $/km automático según unidad
        km_f = (f'=IF(A{r}<>"",IF(C{r}="Pick Up",7,'
                f'IF(C{r}="1.5 Cerrada",9,'
                f'IF(C{r}="2.5",10,'
                f'IF(C{r}="3.5",12,0)))),"")')
        c5 = ws.cell(row=r, column=5, value=km_f)
        c5.fill = fill(AZUL_CLAR)
        c5.border = border_thin()
        c5.alignment = align(h="right")
        c5.number_format = '"$"#,##0'

        # Costo unidad = km * $/km
        c6 = ws.cell(row=r, column=6,
                     value=f'=IF(AND(D{r}<>"",E{r}<>""),D{r}*E{r},"")')
        c6.fill = fill(VERDE_CLAR)
        c6.font = font(bold=True, color=VERDE_OSC)
        c6.border = border_thin()
        c6.number_format = '"$"#,##0.00'
        c6.alignment = align(h="right")

        # Gastos individuales (chofer, ayudantes, maniobras, casetas,
        #                      comidas, hoteles, gasolina+, materiales, otros)
        for col in range(7, 16):
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.number_format = '"$"#,##0.00'
            c.alignment = align(h="right")

        # Total gastos = costo unidad + todos los demás
        c16 = ws.cell(row=r, column=16,
                      value=f'=IF(A{r}<>"",IFERROR(F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r}+N{r}+O{r},0),"")')
        c16.fill = fill(ROJO_CLAR)
        c16.font = font(bold=True, color=ROJO_OSC)
        c16.border = border_thin()
        c16.number_format = '"$"#,##0.00'
        c16.alignment = align(h="right")

        # Precio cobrado (VLOOKUP desde Agenda columna 14)
        c17 = ws.cell(row=r, column=17,
                      value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:N,14,0),0),"")')
        c17.fill = fill(row_bg)
        c17.border = border_thin()
        c17.number_format = '"$"#,##0.00'
        c17.alignment = align(h="right")

        # Utilidad = precio - total gastos
        c18 = ws.cell(row=r, column=18,
                      value=f'=IF(A{r}<>"",Q{r}-P{r},"")')
        c18.fill = fill(VERDE_CLAR)
        c18.font = font(bold=True, color=VERDE_OSC)
        c18.border = border_thin()
        c18.number_format = '"$"#,##0.00'
        c18.alignment = align(h="right")

        # Margen %
        c19 = ws.cell(row=r, column=19,
                      value=f'=IF(AND(A{r}<>"",Q{r}<>0),R{r}/Q{r}*100,"")')
        c19.fill = fill(row_bg)
        c19.border = border_thin()
        c19.number_format = '0.0"%"'
        c19.alignment = align(h="right")

        # Notas
        c20 = ws.cell(row=r, column=20)
        c20.fill = fill(row_bg)
        c20.border = border_thin()
        c20.alignment = align(h="left", wrap=True)

        row_h(ws, r, 18)

    # Formato condicional: margen
    from openpyxl.formatting.rule import ColorScaleRule
    ws.conditional_formatting.add("S4:S203",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF0000",
                       mid_type="num", mid_value=30, mid_color="FFFF00",
                       end_type="num", end_value=60, end_color="00AA00"))

    return ws


# ─── HOJA 3: CONTROL POR UNIDAD ───────────────────────────────────────────────

def build_unidades(wb):
    ws = wb.create_sheet("Unidades")
    ws.sheet_view.showGridLines = False

    merge_header(ws, 1, 1, 9, "MUDANZAS METEPEC — CONTROL POR UNIDAD",
                 bg=AZUL_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 9, "Desempeño · Metas · Ocupación · Rentabilidad",
                 bg=AZUL_MED, size=10)
    row_h(ws, 2, 22)

    # === TABLA MENSUAL ===
    merge_header(ws, 3, 1, 9, "RESUMEN MENSUAL", bg=AZUL_OSC, size=11)
    row_h(ws, 3, 25)

    month_headers = [
        "UNIDAD", "SERVICIOS", "VENTAS $", "COSTOS $",
        "UTILIDAD $", "KM TOTAL", "META MES $", "CUMP. %", "ESTADO"
    ]
    widths = [18, 12, 14, 14, 14, 12, 14, 12, 14]
    for i, (h, w) in enumerate(zip(month_headers, widths), 1):
        header_cell(ws, 4, i, h, bg=AZUL_MED)
        col_w(ws, i, w)
    row_h(ws, 4, 28)

    units = [
        ("Pick Up",      24250),
        ("1.5 Cerrada",  28450),
        ("2.5",          31250),
        ("3.5",          15250),
    ]

    for i, (uname, meta_mes) in enumerate(units, 5):
        row_bg = BLANCO if i % 2 == 0 else GRIS_CLAR
        u = f'"{uname}"'

        # Unidad
        c = ws.cell(row=i, column=1, value=uname)
        c.fill = fill(AZUL_CLAR)
        c.font = font(bold=True, color=AZUL_OSC)
        c.border = border_thin()
        c.alignment = align(h="center")

        # Servicios (COUNTIF en Agenda columna K)
        sv = ws.cell(row=i, column=2,
                     value=f'=COUNTIF(Agenda!K:K,{u})')
        sv.fill = fill(row_bg)
        sv.border = border_thin()
        sv.alignment = align(h="center")
        sv.font = font(bold=True)

        # Ventas (SUMIF en Agenda: unidad en K, precio en N)
        vt = ws.cell(row=i, column=3,
                     value=f'=SUMIF(Agenda!K:K,{u},Agenda!N:N)')
        vt.fill = fill(row_bg)
        vt.border = border_thin()
        vt.number_format = '"$"#,##0.00'
        vt.alignment = align(h="right")
        vt.font = font(bold=True, color=VERDE_OSC)

        # Costos (SUMIF en Costos: unidad en C, total gastos en P)
        ct = ws.cell(row=i, column=4,
                     value=f'=SUMIF(Costos!C:C,{u},Costos!P:P)')
        ct.fill = fill(row_bg)
        ct.border = border_thin()
        ct.number_format = '"$"#,##0.00'
        ct.alignment = align(h="right")
        ct.font = font(color=ROJO_OSC)

        # Utilidad = ventas - costos
        ut = ws.cell(row=i, column=5,
                     value=f'=C{i}-D{i}')
        ut.fill = fill(VERDE_CLAR)
        ut.border = border_thin()
        ut.number_format = '"$"#,##0.00'
        ut.alignment = align(h="right")
        ut.font = font(bold=True, color=VERDE_OSC)

        # KM (SUMIF en Costos: unidad en C, km en D)
        km = ws.cell(row=i, column=6,
                     value=f'=SUMIF(Costos!C:C,{u},Costos!D:D)')
        km.fill = fill(row_bg)
        km.border = border_thin()
        km.number_format = '#,##0.0" km"'
        km.alignment = align(h="right")

        # Meta mensual
        mt = ws.cell(row=i, column=7, value=meta_mes)
        mt.fill = fill(AMBAR_CLAR)
        mt.border = border_thin()
        mt.number_format = '"$"#,##0.00'
        mt.alignment = align(h="right")
        mt.font = font(bold=True, color=AMBAR)

        # Cumplimiento %
        cp = ws.cell(row=i, column=8,
                     value=f'=IF(G{i}>0,C{i}/G{i}*100,0)')
        cp.fill = fill(row_bg)
        cp.border = border_thin()
        cp.number_format = '0.0"%"'
        cp.alignment = align(h="right")
        cp.font = font(bold=True)

        # Estado semáforo
        est = ws.cell(row=i, column=9,
                      value=f'=IF(H{i}>=100,"✅ META CUMPLIDA",'
                            f'IF(H{i}>=70,"⚠️ En camino",'
                            f'"🔴 Atención"))')
        est.fill = fill(row_bg)
        est.border = border_thin()
        est.alignment = align(h="center")

        row_h(ws, i, 22)

    # Totales
    r_tot = 9
    merge_header(ws, r_tot, 1, 1, "TOTAL FLOTA", bg=GRIS_OSC, fg=BLANCO,
                 size=10)
    ws.cell(row=r_tot, column=1).alignment = align(h="center")

    for col, formula in [(2, "=SUM(B5:B8)"), (3, "=SUM(C5:C8)"),
                          (4, "=SUM(D5:D8)"), (5, "=SUM(E5:E8)"),
                          (6, "=SUM(F5:F8)"), (7, "=SUM(G5:G8)")]:
        c = ws.cell(row=r_tot, column=col, value=formula)
        c.fill = fill(GRIS_OSC)
        c.font = font(bold=True, color=BLANCO)
        c.border = border_thin()
        c.alignment = align(h="right" if col > 2 else "center")
        if col in [3, 4, 5, 7]:
            c.number_format = '"$"#,##0.00'
        elif col == 6:
            c.number_format = '#,##0.0" km"'
    row_h(ws, r_tot, 22)

    # === TABLA SEMANAL ===
    r_sem = 11
    merge_header(ws, r_sem, 1, 9, "RESUMEN SEMANAL", bg=AZUL_OSC, size=11)
    row_h(ws, r_sem, 25)

    week_headers = [
        "UNIDAD", "SERVICIOS SEMANA", "VENTAS SEMANA $",
        "META SEMANA $", "CUMP. SEMANAL %", "FALTA $",
        "PROM. x SERVICIO $", "MEJOR SEMANA", "NOTA"
    ]
    for i, h in enumerate(week_headers, 1):
        header_cell(ws, r_sem + 1, i, h, bg=VERDE_OSC)
    row_h(ws, r_sem + 1, 28)

    units_week = [
        ("Pick Up",     5600),
        ("1.5 Cerrada", 6570),
        ("2.5",         7220),
        ("3.5",         3520),
    ]

    for idx, (uname, meta_sem) in enumerate(units_week):
        r = r_sem + 2 + idx
        row_bg = BLANCO if idx % 2 == 0 else GRIS_CLAR
        u = f'"{uname}"'
        row_ref_mes = 5 + idx  # fila en tabla mensual

        c1 = ws.cell(row=r, column=1, value=uname)
        c1.fill = fill(AZUL_CLAR)
        c1.font = font(bold=True, color=AZUL_OSC)
        c1.border = border_thin()
        c1.alignment = align(h="center")

        # Servicios esta semana (fecha en Agenda col C >= hoy-7)
        sv = ws.cell(row=r, column=2,
                     value=f'=COUNTIFS(Agenda!K:K,{u},'
                           f'Agenda!C:C,">="&TODAY()-7,'
                           f'Agenda!C:C,"<="&TODAY())')
        sv.fill = fill(row_bg)
        sv.border = border_thin()
        sv.alignment = align(h="center")
        sv.font = font(bold=True)

        # Ventas esta semana
        vt = ws.cell(row=r, column=3,
                     value=f'=SUMIFS(Agenda!N:N,Agenda!K:K,{u},'
                           f'Agenda!C:C,">="&TODAY()-7,'
                           f'Agenda!C:C,"<="&TODAY())')
        vt.fill = fill(row_bg)
        vt.border = border_thin()
        vt.number_format = '"$"#,##0.00'
        vt.alignment = align(h="right")
        vt.font = font(bold=True, color=VERDE_OSC)

        # Meta semana
        mt = ws.cell(row=r, column=4, value=meta_sem)
        mt.fill = fill(AMBAR_CLAR)
        mt.border = border_thin()
        mt.number_format = '"$"#,##0.00'
        mt.alignment = align(h="right")
        mt.font = font(bold=True, color=AMBAR)

        # Cumplimiento semanal %
        cp = ws.cell(row=r, column=5,
                     value=f'=IF(D{r}>0,C{r}/D{r}*100,0)')
        cp.fill = fill(row_bg)
        cp.border = border_thin()
        cp.number_format = '0.0"%"'
        cp.alignment = align(h="right")
        cp.font = font(bold=True)

        # Falta para meta
        flt = ws.cell(row=r, column=6,
                      value=f'=MAX(0,D{r}-C{r})')
        flt.fill = fill(ROJO_CLAR)
        flt.border = border_thin()
        flt.number_format = '"$"#,##0.00'
        flt.alignment = align(h="right")
        flt.font = font(color=ROJO_OSC)

        # Promedio por servicio
        prm = ws.cell(row=r, column=7,
                      value=f'=IF(B{r}>0,C{r}/B{r},0)')
        prm.fill = fill(row_bg)
        prm.border = border_thin()
        prm.number_format = '"$"#,##0.00'
        prm.alignment = align(h="right")

        ws.cell(row=r, column=8, value="").border = border_thin()
        ws.cell(row=r, column=9, value="").border = border_thin()

        row_h(ws, r, 22)

    return ws


# ─── HOJA 4: FLUJO DE EFECTIVO ────────────────────────────────────────────────

def build_flujo(wb):
    ws = wb.create_sheet("Flujo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    merge_header(ws, 1, 1, 9, "MUDANZAS METEPEC — FLUJO DE EFECTIVO",
                 bg=VERDE_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 9,
                 "Cada movimiento de dinero: entradas · salidas · saldo acumulado",
                 bg=VERDE_MED, size=10)
    row_h(ws, 2, 22)

    headers = [
        ("FECHA", 12), ("CONCEPTO", 35), ("ENTRADA $", 14),
        ("SALIDA $", 14), ("CATEGORÍA", 22), ("UNIDAD", 16),
        ("FOLIO", 12), ("MÉTODO PAGO", 14), ("SALDO ACUM. $", 16),
    ]
    for i, (h, w) in enumerate(headers, 1):
        header_cell(ws, 3, i, h, bg=VERDE_OSC)
        col_w(ws, i, w)
    row_h(ws, 3, 30)

    # Saldo inicial en fila 4
    c_si = ws.cell(row=4, column=2, value="SALDO INICIAL")
    c_si.fill = fill(AZUL_CLAR)
    c_si.font = font(bold=True, color=AZUL_OSC)
    c_si.border = border_thin()
    c_si.alignment = align(h="left")

    for col in [1, 3, 4, 5, 6, 7, 8]:
        c = ws.cell(row=4, column=col)
        c.fill = fill(AZUL_CLAR)
        c.border = border_thin()
        if col in [3, 4]:
            c.number_format = '"$"#,##0.00'
            c.alignment = align(h="right")

    # Saldo inicial calculado
    c_saldo0 = ws.cell(row=4, column=9, value='=IF(C4<>"",C4,0)')
    c_saldo0.fill = fill(VERDE_CLAR)
    c_saldo0.font = font(bold=True, color=VERDE_OSC)
    c_saldo0.border = border_thin()
    c_saldo0.number_format = '"$"#,##0.00'
    c_saldo0.alignment = align(h="right")
    row_h(ws, 4, 22)

    for r in range(5, 505):
        row_bg = BLANCO if r % 2 == 0 else GRIS_CLAR

        c_fecha = ws.cell(row=r, column=1)
        c_fecha.fill = fill(row_bg)
        c_fecha.border = border_thin()
        c_fecha.number_format = "DD/MM/YYYY"
        c_fecha.alignment = align(h="center")

        c_conc = ws.cell(row=r, column=2)
        c_conc.fill = fill(row_bg)
        c_conc.border = border_thin()
        c_conc.alignment = align(h="left")

        # Entrada
        c_ent = ws.cell(row=r, column=3)
        c_ent.fill = fill(VERDE_CLAR)
        c_ent.border = border_thin()
        c_ent.number_format = '"$"#,##0.00'
        c_ent.alignment = align(h="right")

        # Salida
        c_sal = ws.cell(row=r, column=4)
        c_sal.fill = fill(ROJO_CLAR)
        c_sal.border = border_thin()
        c_sal.number_format = '"$"#,##0.00'
        c_sal.alignment = align(h="right")

        for col in [5, 6, 7, 8]:
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="center")

        # Saldo acumulado
        c_sacc = ws.cell(row=r, column=9,
                         value=f'=IF(OR(C{r}<>"",D{r}<>""),'
                               f'I{r-1}+IF(C{r}<>"",C{r},0)-IF(D{r}<>"",D{r},0),"")')
        c_sacc.fill = fill(VERDE_CLAR)
        c_sacc.font = font(bold=True, color=VERDE_OSC)
        c_sacc.border = border_thin()
        c_sacc.number_format = '"$"#,##0.00'
        c_sacc.alignment = align(h="right")

        row_h(ws, r, 18)

    # Dropdowns
    add_dropdown(ws,
        '"Servicio,Gasto operativo,Mensualidad,Publicidad,Administración,Deuda,Nómina,Mantenimiento"',
        "E4:E504")
    add_dropdown(ws,
        '"Pick Up,1.5 Cerrada,2.5,3.5,General"',
        "F4:F504")
    add_dropdown(ws,
        '"Efectivo,Transferencia,Tarjeta,Mixto"',
        "H4:H504")

    # Panel de resumen (columnas 11+)
    col_s = 11
    ws.column_dimensions[get_column_letter(col_s)].width = 24
    ws.column_dimensions[get_column_letter(col_s+1)].width = 16

    resumen_labels = [
        ("RESUMEN FLUJO", None, AZUL_OSC, BLANCO, True),
        ("Total Entradas", "=IFERROR(SUM(C4:C504),0)", VERDE_CLAR, VERDE_OSC, False),
        ("Total Salidas", "=IFERROR(SUM(D4:D504),0)", ROJO_CLAR, ROJO_OSC, False),
        ("Saldo Neto", "=K12-K13", AZUL_CLAR, AZUL_OSC, True),
        ("", None, BLANCO, NEGRO, False),
        ("Por categoría:", None, GRIS_CLAR, GRIS_OSC, True),
        ("Servicios", '=SUMIF(E4:E504,"Servicio",C4:C504)', VERDE_CLAR, VERDE_OSC, False),
        ("G. Operativo", '=SUMIF(E4:E504,"Gasto operativo",D4:D504)', ROJO_CLAR, ROJO_OSC, False),
        ("Nómina", '=SUMIF(E4:E504,"Nómina",D4:D504)', ROJO_CLAR, ROJO_OSC, False),
        ("Publicidad", '=SUMIF(E4:E504,"Publicidad",D4:D504)', AMBAR_CLAR, AMBAR, False),
        ("Mantenimiento", '=SUMIF(E4:E504,"Mantenimiento",D4:D504)', ROJO_CLAR, ROJO_OSC, False),
    ]

    for idx, (label, formula, bg_c, fg_c, bold_) in enumerate(resumen_labels, 11):
        r = idx
        c = ws.cell(row=r, column=col_s, value=label)
        c.fill = fill(bg_c)
        c.font = font(bold=bold_, color=fg_c)
        c.border = border_thin()
        c.alignment = align(h="left")
        row_h(ws, r, 22)

        if formula:
            cv = ws.cell(row=r, column=col_s+1, value=formula)
            cv.fill = fill(bg_c)
            cv.font = font(bold=bold_, color=fg_c)
            cv.border = border_thin()
            cv.number_format = '"$"#,##0.00'
            cv.alignment = align(h="right")

    return ws


# ─── HOJA 5: GASTOS FIJOS ─────────────────────────────────────────────────────

def build_gastos_fijos(wb):
    ws = wb.create_sheet("Gastos Fijos")
    ws.sheet_view.showGridLines = False

    merge_header(ws, 1, 1, 5, "MUDANZAS METEPEC — GASTOS FIJOS",
                 bg=ROJO_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 5,
                 "Gastos fijos mensuales · Punto de equilibrio · Avance de cobertura",
                 bg="C0392B", size=10)
    row_h(ws, 2, 22)

    col_w(ws, 1, 30)
    col_w(ws, 2, 16)
    col_w(ws, 3, 16)
    col_w(ws, 4, 16)
    col_w(ws, 5, 22)

    for i, h in enumerate(["CONCEPTO", "MONTO MES $", "MONTO SEM $",
                            "% DEL TOTAL", "CATEGORÍA"], 1):
        header_cell(ws, 3, i, h, bg=ROJO_OSC)
    row_h(ws, 3, 28)

    gastos = [
        ("Camioneta 2.5",       16000, "Unidad"),
        ("Camioneta 1.5 C",     13200, "Unidad"),
        ("Camioneta Pick Up",    9000, "Unidad"),
        ("Pensión (estacionam)",20000, "Operativo"),
        ("Anuncios / Publicidad", 6000, "Marketing"),
        ("Suscripciones",        5000, "Admin"),
        ("Administración",      30000, "Admin"),
    ]
    total_fijo = sum(g[1] for g in gastos)

    for idx, (concepto, monto, cat) in enumerate(gastos, 4):
        row_bg = BLANCO if idx % 2 == 0 else GRIS_CLAR
        r = idx

        c1 = ws.cell(row=r, column=1, value=concepto)
        c1.fill = fill(row_bg)
        c1.font = font(color=GRIS_OSC)
        c1.border = border_thin()
        c1.alignment = align(h="left")

        c2 = ws.cell(row=r, column=2, value=monto)
        c2.fill = fill(row_bg)
        c2.border = border_thin()
        c2.number_format = '"$"#,##0.00'
        c2.alignment = align(h="right")
        c2.font = font(color=ROJO_OSC)

        c3 = ws.cell(row=r, column=3, value=f'=B{r}/4.33')
        c3.fill = fill(row_bg)
        c3.border = border_thin()
        c3.number_format = '"$"#,##0.00'
        c3.alignment = align(h="right")

        c4 = ws.cell(row=r, column=4, value=f'=B{r}/$B$11*100')
        c4.fill = fill(row_bg)
        c4.border = border_thin()
        c4.number_format = '0.0"%"'
        c4.alignment = align(h="right")

        c5 = ws.cell(row=r, column=5, value=cat)
        c5.fill = fill(row_bg)
        c5.border = border_thin()
        c5.alignment = align(h="center")

        row_h(ws, r, 22)

    # Totales
    r_tot = 11
    ct = ws.cell(row=r_tot, column=1, value="TOTAL MENSUAL")
    ct.fill = fill(ROJO_OSC)
    ct.font = font(bold=True, color=BLANCO, size=11)
    ct.border = border_thin()
    ct.alignment = align(h="left")

    cm = ws.cell(row=r_tot, column=2, value=f'=SUM(B4:B10)')
    cm.fill = fill(ROJO_OSC)
    cm.font = font(bold=True, color=BLANCO, size=11)
    cm.border = border_thin()
    cm.number_format = '"$"#,##0.00'
    cm.alignment = align(h="right")

    cs = ws.cell(row=r_tot, column=3, value=f'=B11/4.33')
    cs.fill = fill(ROJO_OSC)
    cs.font = font(bold=True, color=BLANCO, size=11)
    cs.border = border_thin()
    cs.number_format = '"$"#,##0.00'
    cs.alignment = align(h="right")

    ws.cell(row=r_tot, column=4, value="100%").fill = fill(ROJO_OSC)
    ws.cell(row=r_tot, column=4).font = font(bold=True, color=BLANCO)
    ws.cell(row=r_tot, column=4).border = border_thin()
    ws.cell(row=r_tot, column=4).alignment = align(h="right")
    row_h(ws, r_tot, 26)

    # === PANEL DE COBERTURA ===
    r_panel = 13
    merge_header(ws, r_panel, 1, 5,
                 "AVANCE DE COBERTURA DE GASTOS FIJOS", bg=AZUL_OSC, size=11)
    row_h(ws, r_panel, 28)

    cobertura_items = [
        ("Ventas del mes (total flota)",
         "=IFERROR(SUM(Agenda!N:N),0)", VERDE_CLAR, VERDE_OSC),
        ("Gastos fijos totales",
         "=B11", ROJO_CLAR, ROJO_OSC),
        ("Gastos fijos CUBIERTOS",
         "=MIN(B14,B15)", VERDE_CLAR, VERDE_OSC),
        ("FALTA para cubrir gastos fijos",
         "=MAX(0,B15-B14)", AMBAR_CLAR, AMBAR),
        ("% Cobertura alcanzado",
         "=IF(B15>0,B14/B15*100,0)", AZUL_CLAR, AZUL_OSC),
        ("Ventas necesarias p/equilibrio",
         "=B15", GRIS_CLAR, GRIS_MED),
        ("Margen sobre gastos fijos",
         "=MAX(0,B14-B15)", VERDE_CLAR, VERDE_OSC),
    ]

    for idx, (label, formula, bg_c, fg_c) in enumerate(cobertura_items, r_panel+1):
        r = idx
        c1 = ws.cell(row=r, column=1, value=label)
        c1.fill = fill(bg_c)
        c1.font = font(color=fg_c)
        c1.border = border_thin()
        c1.alignment = align(h="left")

        c2 = ws.cell(row=r, column=2, value=formula)
        c2.fill = fill(bg_c)
        c2.font = font(bold=True, color=fg_c)
        c2.border = border_thin()
        c2.number_format = ('"$"#,##0.00' if "%" not in label else '0.0"%"')
        c2.alignment = align(h="right")

        for col in [3, 4, 5]:
            ws.cell(row=r, column=col).fill = fill(bg_c)
            ws.cell(row=r, column=col).border = border_thin()

        row_h(ws, r, 22)

    # Distribución por unidad
    r_dist = r_panel + len(cobertura_items) + 2
    merge_header(ws, r_dist, 1, 5,
                 "DISTRIBUCIÓN META POR UNIDAD (vs. Gastos Fijos)", bg=AZUL_MED, size=11)
    row_h(ws, r_dist, 25)

    dist_headers = ["UNIDAD", "META MES $", "META SEM $", "% META/GASTO", "VENTAS REALES $"]
    for i, h in enumerate(dist_headers, 1):
        header_cell(ws, r_dist+1, i, h, bg=AZUL_MED)
    row_h(ws, r_dist+1, 26)

    dist_data = [
        ("Pick Up",     24250, 5600),
        ("1.5 Cerrada", 28450, 6570),
        ("2.5",         31250, 7220),
        ("3.5",         15250, 3520),
    ]
    for idx, (u, meta_m, meta_s) in enumerate(dist_data, r_dist+2):
        row_bg = BLANCO if idx % 2 == 0 else GRIS_CLAR
        u_q = f'"{u}"'

        ws.cell(row=idx, column=1, value=u).fill = fill(AZUL_CLAR)
        ws.cell(row=idx, column=1).font = font(bold=True, color=AZUL_OSC)
        ws.cell(row=idx, column=1).border = border_thin()
        ws.cell(row=idx, column=1).alignment = align(h="center")

        ws.cell(row=idx, column=2, value=meta_m).fill = fill(row_bg)
        ws.cell(row=idx, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=idx, column=2).border = border_thin()
        ws.cell(row=idx, column=2).alignment = align(h="right")

        ws.cell(row=idx, column=3, value=meta_s).fill = fill(row_bg)
        ws.cell(row=idx, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=idx, column=3).border = border_thin()
        ws.cell(row=idx, column=3).alignment = align(h="right")

        ws.cell(row=idx, column=4,
                value=f'=B{idx}/$B$11*100').fill = fill(row_bg)
        ws.cell(row=idx, column=4).number_format = '0.0"%"'
        ws.cell(row=idx, column=4).border = border_thin()
        ws.cell(row=idx, column=4).alignment = align(h="right")

        ws.cell(row=idx, column=5,
                value=f'=SUMIF(Agenda!K:K,{u_q},Agenda!N:N)').fill = fill(VERDE_CLAR)
        ws.cell(row=idx, column=5).font = font(bold=True, color=VERDE_OSC)
        ws.cell(row=idx, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=idx, column=5).border = border_thin()
        ws.cell(row=idx, column=5).alignment = align(h="right")

        row_h(ws, idx, 22)

    return ws


# ─── HOJA 6: CLIENTES Y SEGUIMIENTO ──────────────────────────────────────────

def build_clientes(wb):
    ws = wb.create_sheet("Clientes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    merge_header(ws, 1, 1, 16, "MUDANZAS METEPEC — CLIENTES Y SEGUIMIENTO",
                 bg=AZUL_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 16,
                 "CRM básico: contacto → cotización → cierre → reseña → recurrencia",
                 bg=AZUL_MED, size=10)
    row_h(ws, 2, 22)

    headers = [
        ("NOMBRE", 22), ("TELÉFONO", 14), ("F.CONTACTO", 12),
        ("FUENTE", 16), ("SERVICIO SOL.", 22), ("COTIZACIÓN $", 13),
        ("CERRÓ", 9), ("MOTIVO NO CERRÓ", 22), ("F.SEGUIMIENTO", 13),
        ("RESEÑA SOL.", 11), ("RESEÑA REC.", 11), ("RECURRENTE", 11),
        ("VALOR CLIENTE $", 14), ("SERVICIOS TOTAL", 12),
        ("ÚLTIMO SERVICIO", 13), ("OBSERVACIONES", 30),
    ]
    for i, (h, w) in enumerate(headers, 1):
        header_cell(ws, 3, i, h)
        col_w(ws, i, w)
    row_h(ws, 3, 30)

    for r in range(4, 304):
        row_bg = BLANCO if r % 2 == 0 else GRIS_CLAR

        for col in [1, 2, 5, 8, 16]:
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="left", wrap=(col in [8, 16]))

        for col in [3, 9, 15]:
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.number_format = "DD/MM/YYYY"
            c.alignment = align(h="center")

        for col in [4, 6, 7, 10, 11, 12]:
            c = ws.cell(row=r, column=col)
            c.fill = fill(row_bg)
            c.border = border_thin()
            c.alignment = align(h="center")

        # Cotización $
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6).alignment = align(h="right")

        # Valor cliente (SUMIF en Agenda por nombre)
        vc = ws.cell(row=r, column=13,
                     value=f'=IF(A{r}<>"",IFERROR(SUMIF(Agenda!E:E,A{r},Agenda!N:N),0),"")')
        vc.fill = fill(VERDE_CLAR)
        vc.font = font(bold=True, color=VERDE_OSC)
        vc.border = border_thin()
        vc.number_format = '"$"#,##0.00'
        vc.alignment = align(h="right")

        # Servicios totales
        st = ws.cell(row=r, column=14,
                     value=f'=IF(A{r}<>"",IFERROR(COUNTIF(Agenda!E:E,A{r}),0),"")')
        st.fill = fill(AZUL_CLAR)
        st.font = font(bold=True, color=AZUL_OSC)
        st.border = border_thin()
        st.alignment = align(h="center")

        # Último servicio (MAXIFS)
        us = ws.cell(row=r, column=15,
                     value=f'=IF(A{r}<>"",IFERROR(MAXIFS(Agenda!C:C,Agenda!E:E,A{r}),""),"")')
        us.fill = fill(row_bg)
        us.border = border_thin()
        us.number_format = "DD/MM/YYYY"
        us.alignment = align(h="center")

        row_h(ws, r, 18)

    # Dropdowns
    add_dropdown(ws,
        '"Saleads,Facebook,WhatsApp,Recomendación,Google,Otro"',
        "D4:D303")
    add_dropdown(ws, '"Sí,No"', "G4:G303")
    add_dropdown(ws, '"Sí,No"', "J4:J303")
    add_dropdown(ws, '"Sí,No"', "K4:K303")
    add_dropdown(ws, '"Sí,No"', "L4:L303")

    # Colores condicionales
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add("G4:G303",
        CellIsRule(operator="equal", formula=['"Sí"'],
                   fill=PatternFill("solid", fgColor=VERDE_CLAR),
                   font=Font(color=VERDE_OSC, bold=True)))
    ws.conditional_formatting.add("G4:G303",
        CellIsRule(operator="equal", formula=['"No"'],
                   fill=PatternFill("solid", fgColor=ROJO_CLAR),
                   font=Font(color=ROJO_OSC, bold=True)))

    # Panel métricas CRM
    col_m = 18
    ws.column_dimensions[get_column_letter(col_m)].width = 28
    ws.column_dimensions[get_column_letter(col_m+1)].width = 16

    crm_items = [
        ("MÉTRICAS CRM", None, AZUL_OSC, BLANCO),
        ("Total contactos", "=COUNTA(A4:A303)", AZUL_CLAR, AZUL_OSC),
        ("Total cerrados", '=COUNTIF(G4:G303,"Sí")', VERDE_CLAR, VERDE_OSC),
        ("No cerrados", '=COUNTIF(G4:G303,"No")', ROJO_CLAR, ROJO_OSC),
        ("Tasa de cierre %",
         '=IF(COUNTA(A4:A303)>0,COUNTIF(G4:G303,"Sí")/COUNTA(A4:A303)*100,0)',
         AZUL_CLAR, AZUL_OSC),
        ("Reseñas recibidas", '=COUNTIF(K4:K303,"Sí")', VERDE_CLAR, VERDE_OSC),
        ("Clientes recurrentes", '=COUNTIF(L4:L303,"Sí")', VERDE_CLAR, VERDE_OSC),
        ("Fuente: Facebook", '=COUNTIF(D4:D303,"Facebook")', GRIS_CLAR, GRIS_OSC),
        ("Fuente: WhatsApp", '=COUNTIF(D4:D303,"WhatsApp")', GRIS_CLAR, GRIS_OSC),
        ("Fuente: Recomendación", '=COUNTIF(D4:D303,"Recomendación")', GRIS_CLAR, GRIS_OSC),
        ("Fuente: Saleads", '=COUNTIF(D4:D303,"Saleads")', GRIS_CLAR, GRIS_OSC),
    ]

    for idx, (label, formula, bg_c, fg_c) in enumerate(crm_items, 3):
        r = idx
        c1 = ws.cell(row=r, column=col_m, value=label)
        c1.fill = fill(bg_c)
        c1.font = font(bold=(formula is None), color=fg_c)
        c1.border = border_thin()
        c1.alignment = align(h="left")
        row_h(ws, r, 22)

        if formula:
            c2 = ws.cell(row=r, column=col_m+1, value=formula)
            c2.fill = fill(bg_c)
            c2.font = font(bold=True, color=fg_c)
            c2.border = border_thin()
            c2.number_format = ('0.0"%"' if "%" in label else '#,##0')
            c2.alignment = align(h="right")

    return ws


# ─── HOJA 7: DASHBOARD ────────────────────────────────────────────────────────

def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    merge_header(ws, 1, 1, 12,
                 "MUDANZAS METEPEC — DASHBOARD OPERATIVO",
                 bg=AZUL_OSC, size=16)
    row_h(ws, 1, 45)

    # Fecha actual
    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1,
                 value='="Actualizado: "&TEXT(TODAY(),"DD/MM/YYYY")')
    c2.fill = fill(AZUL_MED)
    c2.font = font(bold=True, size=11, color=BLANCO)
    c2.alignment = align(h="right")
    row_h(ws, 2, 22)

    col_w(ws, 1, 2)
    for c in range(2, 13):
        col_w(ws, c, 16)

    # ── SECCIÓN: SERVICIOS ──
    def kpi_block(ws, row, col, title, formula, bg_title, bg_value, fg_v,
                  num_fmt=None, row_h_val=40):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+1)
        t = ws.cell(row=row, column=col, value=title)
        t.fill = fill(bg_title)
        t.font = font(bold=True, size=9, color=BLANCO)
        t.alignment = align(h="center")
        t.border = border_thin()

        ws.merge_cells(start_row=row+1, start_column=col,
                       end_row=row+1, end_column=col+1)
        v = ws.cell(row=row+1, column=col, value=formula)
        v.fill = fill(bg_value)
        v.font = font(bold=True, size=18, color=fg_v)
        v.alignment = align(h="center")
        v.border = border_thin()
        if num_fmt:
            v.number_format = num_fmt
        row_h(ws, row+1, row_h_val)
        return v

    r = 4
    row_h(ws, r, 22)
    row_h(ws, r+1, 45)

    # Fila 1 de KPIs - Servicios
    ws.merge_cells(f"B{r}:M{r}")
    sec = ws.cell(row=r, column=2, value="SERVICIOS")
    sec.fill = fill(GRIS_OSC)
    sec.font = font(bold=True, color=BLANCO, size=11)
    sec.alignment = align(h="center")
    sec.border = border_thin()
    r += 1

    kpi_data_servicios = [
        ("Servicios HOY", f'=COUNTIF(Agenda!C:C,TODAY())', AZUL_OSC, AZUL_CLAR, AZUL_OSC),
        ("Esta SEMANA", f'=COUNTIFS(Agenda!C:C,">="&TODAY()-WEEKDAY(TODAY(),2)+1,Agenda!C:C,"<="&TODAY()-WEEKDAY(TODAY(),2)+7)', AZUL_MED, AZUL_CLAR, AZUL_OSC),
        ("Este MES", f'=COUNTIFS(Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))', AZUL_OSC, AZUL_CLAR, AZUL_OSC),
        ("PENDIENTES", f'=COUNTIF(Agenda!Q:Q,"Pendiente")', AMBAR, AMBAR_CLAR, AMBAR),
        ("CONFIRMADOS", f'=COUNTIF(Agenda!Q:Q,"Confirmado")', VERDE_MED, VERDE_CLAR, VERDE_OSC),
        ("TERMINADOS mes", f'=COUNTIFS(Agenda!Q:Q,"Terminado",Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))', VERDE_OSC, VERDE_CLAR, VERDE_OSC),
    ]

    col_pos = 2
    for title, formula, bg_t, bg_v, fg_v in kpi_data_servicios:
        kpi_block(ws, r, col_pos, title, formula, bg_t, bg_v, fg_v)
        col_pos += 2

    r += 2
    row_h(ws, r, 12)  # espaciador
    r += 1

    # Fila 2 - Finanzas
    ws.merge_cells(f"B{r}:M{r}")
    sec2 = ws.cell(row=r, column=2, value="FINANZAS DEL MES")
    sec2.fill = fill(VERDE_OSC)
    sec2.font = font(bold=True, color=BLANCO, size=11)
    sec2.alignment = align(h="center")
    sec2.border = border_thin()
    row_h(ws, r, 22)
    r += 1
    row_h(ws, r+1, 45)

    kpi_finanzas = [
        ("VENTAS MES",
         f'=IFERROR(SUMIFS(Agenda!N:N,Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0)),0)',
         VERDE_OSC, VERDE_CLAR, VERDE_OSC, '"$"#,##0'),
        ("COSTOS MES",
         f'=IFERROR(SUMIFS(Costos!P:P,Costos!P:P,">"&0),0)',
         ROJO_OSC, ROJO_CLAR, ROJO_OSC, '"$"#,##0'),
        ("UTILIDAD MES",
         f'=IFERROR(SUMIFS(Agenda!N:N,Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))-SUMIF(Costos!P:P,">"&0,Costos!P:P),0)',
         VERDE_OSC, VERDE_CLAR, VERDE_OSC, '"$"#,##0'),
        ("GASTOS FIJOS",
         f'=IFERROR(\'Gastos Fijos\'!B11,0)',
         ROJO_OSC, ROJO_CLAR, ROJO_OSC, '"$"#,##0'),
        ("FLUJO NETO",
         f'=IFERROR(SUM(Flujo!C4:C504)-SUM(Flujo!D4:D504),0)',
         AZUL_OSC, AZUL_CLAR, AZUL_OSC, '"$"#,##0'),
        ("MARGEN PROM %",
         f'=IFERROR(AVERAGE(Costos!S4:S203),0)',
         AZUL_MED, AZUL_CLAR, AZUL_OSC, '0.0"%"'),
    ]

    col_pos = 2
    for title, formula, bg_t, bg_v, fg_v, nfmt in kpi_finanzas:
        kpi_block(ws, r, col_pos, title, formula, bg_t, bg_v, fg_v,
                  num_fmt=nfmt)
        col_pos += 2

    r += 2
    row_h(ws, r, 12)
    r += 1

    # Fila 3 - Unidades
    ws.merge_cells(f"B{r}:M{r}")
    sec3 = ws.cell(row=r, column=2, value="VENTAS POR UNIDAD — MES ACTUAL")
    sec3.fill = fill(AZUL_OSC)
    sec3.font = font(bold=True, color=BLANCO, size=11)
    sec3.alignment = align(h="center")
    sec3.border = border_thin()
    row_h(ws, r, 22)
    r += 1

    unit_headers = ["UNIDAD", "SERVICIOS", "VENTAS $", "META MES $",
                    "CUMP %", "FALTA $", "UTILIDAD $"]
    for i, h in enumerate(unit_headers, 2):
        header_cell(ws, r, i, h, bg=AZUL_MED)
        row_h(ws, r, 26)

    r += 1
    units_dash = [
        ("Pick Up",     24250, AMBAR_CLAR),
        ("1.5 Cerrada", 28450, VERDE_CLAR),
        ("2.5",         31250, AZUL_CLAR),
        ("3.5",         15250, GRIS_CLAR),
    ]
    for uname, meta, bg_u in units_dash:
        u_q = f'"{uname}"'
        row_bg = bg_u

        ws.cell(row=r, column=2, value=uname).fill = fill(row_bg)
        ws.cell(row=r, column=2).font = font(bold=True, color=GRIS_OSC)
        ws.cell(row=r, column=2).border = border_thin()
        ws.cell(row=r, column=2).alignment = align(h="center")

        # Servicios
        sv = ws.cell(row=r, column=3,
                     value=f'=COUNTIFS(Agenda!K:K,{u_q},'
                           f'Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
                           f'Agenda!C:C,"<="&EOMONTH(TODAY(),0))')
        sv.fill = fill(row_bg)
        sv.border = border_thin()
        sv.alignment = align(h="center")
        sv.font = font(bold=True)

        # Ventas
        vt = ws.cell(row=r, column=4,
                     value=f'=SUMIFS(Agenda!N:N,Agenda!K:K,{u_q},'
                           f'Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
                           f'Agenda!C:C,"<="&EOMONTH(TODAY(),0))')
        vt.fill = fill(row_bg)
        vt.border = border_thin()
        vt.number_format = '"$"#,##0.00'
        vt.alignment = align(h="right")
        vt.font = font(bold=True, color=VERDE_OSC)

        # Meta
        mt = ws.cell(row=r, column=5, value=meta)
        mt.fill = fill(row_bg)
        mt.border = border_thin()
        mt.number_format = '"$"#,##0.00'
        mt.alignment = align(h="right")
        mt.font = font(color=AMBAR)

        # Cumplimiento
        cp = ws.cell(row=r, column=6,
                     value=f'=IF(E{r}>0,D{r}/E{r}*100,0)')
        cp.fill = fill(row_bg)
        cp.border = border_thin()
        cp.number_format = '0.0"%"'
        cp.alignment = align(h="right")
        cp.font = font(bold=True)

        # Falta
        fl = ws.cell(row=r, column=7,
                     value=f'=MAX(0,E{r}-D{r})')
        fl.fill = fill(row_bg)
        fl.border = border_thin()
        fl.number_format = '"$"#,##0.00'
        fl.alignment = align(h="right")
        fl.font = font(color=ROJO_OSC)

        # Utilidad
        ut = ws.cell(row=r, column=8,
                     value=f'=SUMIF(Costos!C:C,{u_q},Costos!R:R)')
        ut.fill = fill(row_bg)
        ut.border = border_thin()
        ut.number_format = '"$"#,##0.00'
        ut.alignment = align(h="right")
        ut.font = font(bold=True, color=VERDE_OSC)

        row_h(ws, r, 22)
        r += 1

    # Fila 4 - CRM
    r += 1
    ws.merge_cells(f"B{r}:M{r}")
    sec4 = ws.cell(row=r, column=2, value="CRM — CONVERSIÓN DE CLIENTES")
    sec4.fill = fill(GRIS_OSC)
    sec4.font = font(bold=True, color=BLANCO, size=11)
    sec4.alignment = align(h="center")
    sec4.border = border_thin()
    row_h(ws, r, 22)
    r += 1
    row_h(ws, r+1, 45)

    kpi_crm = [
        ("CONTACTOS", f'=COUNTA(Clientes!A4:A303)', AZUL_OSC, AZUL_CLAR, AZUL_OSC, None),
        ("COTIZACIONES", f'=COUNTA(Clientes!F4:F303)', AZUL_MED, AZUL_CLAR, AZUL_OSC, None),
        ("CERRADOS",
         f'=COUNTIF(Clientes!G4:G303,"Sí")', VERDE_MED, VERDE_CLAR, VERDE_OSC, None),
        ("TASA CIERRE",
         f'=IF(COUNTA(Clientes!A4:A303)>0,COUNTIF(Clientes!G4:G303,"Sí")/COUNTA(Clientes!A4:A303)*100,0)',
         VERDE_OSC, VERDE_CLAR, VERDE_OSC, '0.0"%"'),
        ("RESEÑAS",
         f'=COUNTIF(Clientes!K4:K303,"Sí")', AMBAR, AMBAR_CLAR, AMBAR, None),
        ("RECURRENTES",
         f'=COUNTIF(Clientes!L4:L303,"Sí")', VERDE_OSC, VERDE_CLAR, VERDE_OSC, None),
    ]

    col_pos = 2
    for title, formula, bg_t, bg_v, fg_v, nfmt in kpi_crm:
        kpi_block(ws, r, col_pos, title, formula, bg_t, bg_v, fg_v,
                  num_fmt=nfmt)
        col_pos += 2

    r += 2
    row_h(ws, r, 12)
    r += 1

    # Gastos fijos cobertura
    ws.merge_cells(f"B{r}:G{r}")
    sec5 = ws.cell(row=r, column=2, value="COBERTURA GASTOS FIJOS")
    sec5.fill = fill(ROJO_OSC)
    sec5.font = font(bold=True, color=BLANCO, size=11)
    sec5.alignment = align(h="center")
    sec5.border = border_thin()

    ws.merge_cells(f"H{r}:M{r}")
    sec6 = ws.cell(row=r, column=8, value="MEJOR UNIDAD DEL MES")
    sec6.fill = fill(VERDE_OSC)
    sec6.font = font(bold=True, color=BLANCO, size=11)
    sec6.alignment = align(h="center")
    sec6.border = border_thin()
    row_h(ws, r, 22)
    r += 1

    # Cobertura - fila valor
    ws.merge_cells(f"B{r}:G{r}")
    cob = ws.cell(row=r, column=2,
                  value=f'="Ventas: $"&TEXT(IFERROR(SUM(Agenda!N:N),0),"#,##0")&" / Meta: $"&TEXT(IFERROR(\'Gastos Fijos\'!B11,0),"#,##0")&" ("&TEXT(IFERROR(SUM(Agenda!N:N)/\'Gastos Fijos\'!B11*100,0),"0.0")&"%)"')
    cob.fill = fill(ROJO_CLAR)
    cob.font = font(bold=True, size=12, color=ROJO_OSC)
    cob.alignment = align(h="center")
    cob.border = border_thin()
    row_h(ws, r, 38)

    ws.merge_cells(f"H{r}:M{r}")
    mejor = ws.cell(row=r, column=8,
                    value=f'=IFERROR(INDEX({"{{"}\"Pick Up\",\"1.5 Cerrada\",\"2.5\",\"3.5\"{"}}"}'
                          f',MATCH(MAX(SUMIF(Agenda!K:K,"Pick Up",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"1.5 Cerrada",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"2.5",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"3.5",Agenda!N:N)),'
                          f'{{SUMIF(Agenda!K:K,"Pick Up",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"1.5 Cerrada",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"2.5",Agenda!N:N),'
                          f'SUMIF(Agenda!K:K,"3.5",Agenda!N:N)}},'
                          f'0)),"—")')
    mejor.fill = fill(VERDE_CLAR)
    mejor.font = font(bold=True, size=16, color=VERDE_OSC)
    mejor.alignment = align(h="center")
    mejor.border = border_thin()

    return ws


# ─── HOJA 8: INSTRUCCIONES ────────────────────────────────────────────────────

def build_instrucciones(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_view.showGridLines = False

    col_w(ws, 1, 4)
    col_w(ws, 2, 60)
    col_w(ws, 3, 40)

    merge_header(ws, 1, 1, 3,
                 "MUDANZAS METEPEC — GUÍA DE USO DEL SISTEMA",
                 bg=AZUL_OSC, size=14)
    row_h(ws, 1, 40)

    instrucciones = [
        ("", BLANCO),
        ("FLUJO DE TRABAJO DIARIO:", AZUL_OSC),
        ("1. AGENDA → Registra cada servicio desde WhatsApp. El folio se genera solo.", BLANCO),
        ("2. AGENDA → Llena: cliente, fecha, hora, tipo, origen, destino, unidad, precio cotizado, anticipo.", GRIS_CLAR),
        ("3. AGENDA → Cambia el Estado conforme avanza: Pendiente → Confirmado → En proceso → Terminado.", BLANCO),
        ("4. COSTOS → Copia el Folio de la Agenda. El cliente y la unidad se llenan solos.", GRIS_CLAR),
        ("5. COSTOS → Llena kilómetros reales, chofer, ayudantes, maniobras, casetas, comidas, etc.", BLANCO),
        ("6. COSTOS → La utilidad y el margen se calculan automáticamente.", GRIS_CLAR),
        ("7. FLUJO → Registra cada entrada de dinero (anticipo, saldo) y cada gasto del día.", BLANCO),
        ("8. CLIENTES → Registra cada contacto nuevo desde WhatsApp o redes. Llena si cerró.", GRIS_CLAR),
        ("9. DASHBOARD → Revisa cada mañana el estado del negocio en 30 segundos.", BLANCO),
        ("", BLANCO),
        ("METAS MENSUALES:", AZUL_OSC),
        ("Pick Up: $24,250 | 1.5 Cerrada: $28,450 | 2.5: $31,250 | 3.5: $15,250", AMBAR_CLAR),
        ("Total flota: $99,200 (igual que gastos fijos — ahí empieza la utilidad)", AMBAR_CLAR),
        ("", BLANCO),
        ("METAS SEMANALES:", AZUL_OSC),
        ("Pick Up: $5,600 | 1.5 Cerrada: $6,570 | 2.5: $7,220 | 3.5: $3,520", VERDE_CLAR),
        ("", BLANCO),
        ("COSTO POR KM (para cálculo automático en hoja Costos):", AZUL_OSC),
        ("Pick Up: $7/km  |  1.5 Cerrada: $9/km  |  2.5: $10/km  |  3.5: $12/km", GRIS_CLAR),
        ("", BLANCO),
        ("TIPS IMPORTANTES:", ROJO_OSC),
        ("- El folio conecta TODO: Agenda → Costos → Flujo. No lo cambies.", ROJO_CLAR),
        ("- Usa los dropdowns (listas) para Estado, Tipo, Unidad — así los cálculos funcionan.", ROJO_CLAR),
        ("- Agrega el anticipo el mismo día que lo recibes en Flujo de Efectivo.", ROJO_CLAR),
        ("- Revisa Unidades cada lunes para ver qué unidad va atrasada en su meta.", ROJO_CLAR),
        ("- Dashboard se actualiza solo cada vez que abres el archivo.", ROJO_CLAR),
        ("", BLANCO),
        ("GASTOS FIJOS MENSUALES TOTALES: $99,200", VERDE_OSC),
        ("Estos se cubren con las ventas de todas las unidades combinadas.", VERDE_CLAR),
        ("Todo lo que supere $99,200 en ventas es utilidad real del negocio.", VERDE_CLAR),
    ]

    for idx, (text, bg_c) in enumerate(instrucciones, 2):
        r = idx
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=text)
        is_header = bg_c in [AZUL_OSC, VERDE_OSC, ROJO_OSC]
        c.fill = fill(bg_c if bg_c != BLANCO else BLANCO)
        c.font = font(
            bold=is_header,
            size=11 if is_header else 10,
            color=BLANCO if is_header else (GRIS_OSC if bg_c == GRIS_CLAR else NEGRO),
        )
        c.alignment = align(h="left", v="center")
        c.border = border_thin()
        row_h(ws, r, 22 if not is_header else 28)

        ws.cell(row=r, column=1).fill = fill(AZUL_OSC if is_header else AZUL_MED)
        ws.cell(row=r, column=1).border = border_thin()

    return ws


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    print("Construyendo hojas...")
    build_dashboard(wb)
    build_agenda(wb)
    build_costos(wb)
    build_unidades(wb)
    build_flujo(wb)
    build_gastos_fijos(wb)
    build_clientes(wb)
    build_instrucciones(wb)

    print("Ajustando propiedades del libro...")
    wb.active = wb["Dashboard"]

    out = "MudanzasMetepec_Sistema.xlsx"
    wb.save(out)
    print(f"\n✅  Archivo generado: {out}")
    print("    Ábrelo en Excel o Google Sheets.")
    print("    Para Google Sheets: Archivo → Importar → Subir el .xlsx")


if __name__ == "__main__":
    main()
