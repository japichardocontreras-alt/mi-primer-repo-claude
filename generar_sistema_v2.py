"""
Generador v2 - Sistema Administrativo Mudanzas Metepec
BACKUP v1 conservado en: MudanzasMetepec_Sistema_BACKUP_v1.xlsx
Genera: MudanzasMetepec_Sistema_v2.xlsx

Cambios sobre v1 (sin tocar nada existente, solo agregados):
  - Agenda: 15 columnas nuevas (cols 23-37)
  - Unidades: Expediente técnico + Indicadores + Ocupación
  - Nueva hoja: Mantenimiento (historial por unidad)
  - Dashboard: bloque adicional de KPIs operativos
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ── PALETA (idéntica a v1) ────────────────────────────────────────────────────
AZUL_OSC   = "1A3C5E"
AZUL_MED   = "2E6DA4"
AZUL_CLAR  = "D6E4F0"
VERDE_OSC  = "1E6B3C"
VERDE_MED  = "27AE60"
VERDE_CLAR = "D5F5E3"
ROJO_OSC   = "922B21"
ROJO_CLAR  = "FADBD8"
AMBAR      = "F39C12"
AMBAR_CLAR = "FEF9E7"
GRIS_OSC   = "2C3E50"
GRIS_MED   = "7F8C8D"
GRIS_CLAR  = "F2F3F4"
BLANCO     = "FFFFFF"
NEGRO      = "000000"

# ── HELPERS (idénticos a v1) ──────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, size=10, color=NEGRO, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def border_medium():
    s = Side(style="medium", color=AZUL_MED)
    return Border(left=s, right=s, top=s, bottom=s)
def col_w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width
def row_h(ws, row, height):
    ws.row_dimensions[row].height = height
def header_cell(ws, row, col, text, bg=AZUL_OSC, fg=BLANCO, size=10, bold=True, h_align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold=bold, size=size, color=fg)
    c.alignment = align(h=h_align, v="center", wrap=True)
    c.border = border_thin(); return c
def merge_header(ws, row, c1, c2, text, bg=AZUL_OSC, fg=BLANCO, size=12):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fill(bg); c.font = font(bold=True, size=size, color=fg)
    c.alignment = align(h="center", v="center"); c.border = border_medium(); return c
def add_dropdown(ws, formula, sqref):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = sqref; ws.add_data_validation(dv)
def money_cell(ws, r, col, formula=None, value=None, bg=BLANCO, bold=False, fg=GRIS_OSC):
    v = formula if formula is not None else value
    c = ws.cell(row=r, column=col, value=v)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg)
    c.border = border_thin(); c.number_format = '"$"#,##0.00'
    c.alignment = align(h="right"); return c

# ── AGENDA (v1 íntegra + 15 columnas nuevas al final) ────────────────────────
def build_agenda(wb):
    ws = wb.create_sheet("Agenda")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Títulos extendidos a col 37 (v1 tenía 22)
    merge_header(ws, 1, 1, 37, "MUDANZAS METEPEC — AGENDA DE SERVICIOS", bg=AZUL_OSC, size=14)
    row_h(ws, 1, 35)
    merge_header(ws, 2, 1, 37, "Registro de todos los servicios: WhatsApp → Cobro → Análisis", bg=AZUL_MED, size=10)
    row_h(ws, 2, 22)

    # ── Encabezados v1 (cols 1-22, sin cambio) ──
    v1_headers = [
        ("FOLIO",8),("F.REGISTRO",12),("F.SERVICIO",12),("HORA",8),
        ("CLIENTE",20),("TELÉFONO",13),("TIPO",16),("ORIGEN",20),
        ("DESTINO",20),("KM BASE",9),("UNIDAD",16),("PERSONAL",22),
        ("DESCRIPCIÓN",28),("COTIZADO",11),("ANTICIPO",11),("SALDO",11),
        ("ESTADO",14),("PAGO",13),("FACTURA",9),("CANT.PERSONAS",10),
        ("FOLIO COSTO",10),("OBSERVACIONES",30),
    ]
    for i,(h,w) in enumerate(v1_headers,1):
        header_cell(ws,3,i,h); col_w(ws,i,w)

    # ── Encabezados nuevos (cols 23-37) ──
    new_headers = [
        ("PRIORIDAD",12),("HORA SALIDA",11),("HORA LLEGADA",11),
        ("HORA TÉRMINO",12),("T.ESTIMADO hrs",12),("T.REAL hrs",11),
        ("FOTOS",9),("VIDEOS",9),("RESEÑA SOL.",11),("RESEÑA REC.",11),
        ("PERM.PUBLICAR",12),("FUENTE",16),("TIPO CLIENTE",14),("OBS.CLIENTE",30),
        ("CANCELACIÓN",18),
    ]
    for i,(h,w) in enumerate(new_headers, 23):
        header_cell(ws, 3, i, h, bg=AZUL_MED)
        col_w(ws, i, w)
    row_h(ws, 3, 30)

    # ── Filas de datos (200 filas = v1 íntegra + nuevas cols) ──
    for r in range(4, 204):
        rb = BLANCO if r%2==0 else GRIS_CLAR

        # --- cols v1 (copiadas exactas de v1) ---
        c=ws.cell(row=r,column=1,value=f'=IF(E{r}<>"","MUD-"&TEXT(ROW()-3,"000"),"")')
        c.fill=fill(AZUL_CLAR); c.font=font(bold=True,color=AZUL_OSC)
        c.alignment=align(h="center"); c.border=border_thin()

        c2=ws.cell(row=r,column=2,value=f'=IF(E{r}<>"",TODAY(),"")')
        c2.fill=fill(rb); c2.font=font(color=GRIS_OSC)
        c2.alignment=align(h="center"); c2.border=border_thin()
        c2.number_format="DD/MM/YYYY"

        for col in [3]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.number_format="DD/MM/YYYY"; c.alignment=align(h="center")
        for col in [4]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.number_format="HH:MM"; c.alignment=align(h="center")
        for col in [5,6,8,9,12,13,22]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.alignment=align(h="left",wrap=(col in [12,13,22]))
        for col in [10,20]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.alignment=align(h="center"); c.number_format="#,##0.0"
        for col in [14,15]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.number_format='"$"#,##0.00'; c.alignment=align(h="right")

        cs=ws.cell(row=r,column=16,value=f'=IF(N{r}<>"",N{r}-O{r},"")')
        cs.fill=fill(AMBAR_CLAR); cs.font=font(bold=True,color=GRIS_OSC)
        cs.border=border_thin(); cs.number_format='"$"#,##0.00'
        cs.alignment=align(h="right")

        for col in [7,11,17,18,19]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.alignment=align(h="center")

        cf=ws.cell(row=r,column=21,value=f'=IF(E{r}<>"","MUD-"&TEXT(ROW()-3,"000"),"")')
        cf.fill=fill(AZUL_CLAR); cf.font=font(color=AZUL_OSC)
        cf.border=border_thin(); cf.alignment=align(h="center")

        # --- cols nuevas (23-37) ---
        # 23 PRIORIDAD
        c=ws.cell(row=r,column=23); c.fill=fill(rb); c.border=border_thin(); c.alignment=align(h="center")

        # 24 HORA SALIDA
        c=ws.cell(row=r,column=24); c.fill=fill(rb); c.border=border_thin()
        c.number_format="HH:MM"; c.alignment=align(h="center")

        # 25 HORA LLEGADA
        c=ws.cell(row=r,column=25); c.fill=fill(rb); c.border=border_thin()
        c.number_format="HH:MM"; c.alignment=align(h="center")

        # 26 HORA TÉRMINO
        c=ws.cell(row=r,column=26); c.fill=fill(rb); c.border=border_thin()
        c.number_format="HH:MM"; c.alignment=align(h="center")

        # 27 TIEMPO ESTIMADO
        c=ws.cell(row=r,column=27); c.fill=fill(rb); c.border=border_thin()
        c.number_format='0.0" hrs"'; c.alignment=align(h="center")

        # 28 TIEMPO REAL = HORA TÉRMINO - HORA SALIDA (en horas)
        c=ws.cell(row=r,column=28,
            value=f'=IF(AND(X{r}<>"",Z{r}<>""),IFERROR((Z{r}-X{r})*24,""),"")' )
        c.fill=fill(AZUL_CLAR); c.font=font(color=AZUL_OSC)
        c.border=border_thin(); c.number_format='0.0" hrs"'; c.alignment=align(h="center")

        # 29-33 Sí/No
        for col in [29,30,31,32,33]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.alignment=align(h="center")

        # 34 FUENTE
        c=ws.cell(row=r,column=34); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="center")

        # 35 TIPO CLIENTE
        c=ws.cell(row=r,column=35); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="center")

        # 36 OBS CLIENTE
        c=ws.cell(row=r,column=36); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="left",wrap=True)

        # 37 MOTIVO CANCELACIÓN
        c=ws.cell(row=r,column=37); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="left",wrap=True)

        row_h(ws,r,18)

    # ── Dropdowns v1 (intactos) ──
    add_dropdown(ws,'"Mudanza local,Flete,Renta de unidad,Foráneo,Maniobra especial"',"G4:G203")
    add_dropdown(ws,'"Pick Up,1.5 Cerrada,2.5,3.5,Varias unidades"',"K4:K203")
    add_dropdown(ws,'"Pendiente,Confirmado,En proceso,Terminado,Cancelado"',"Q4:Q203")
    add_dropdown(ws,'"Efectivo,Transferencia,Tarjeta,Mixto"',"R4:R203")
    add_dropdown(ws,'"Sí,No"',"S4:S203")

    # ── Dropdowns nuevos ──
    add_dropdown(ws,'"Alta,Media,Baja"',"W4:W203")          # col 23
    add_dropdown(ws,'"Sí,No"',"AC4:AC203")                   # col 29 fotos
    add_dropdown(ws,'"Sí,No"',"AD4:AD203")                   # col 30 videos
    add_dropdown(ws,'"Sí,No"',"AE4:AE203")                   # col 31 reseña sol
    add_dropdown(ws,'"Sí,No"',"AF4:AF203")                   # col 32 reseña rec
    add_dropdown(ws,'"Sí,No"',"AG4:AG203")                   # col 33 perm publicar
    add_dropdown(ws,'"Saleads,Facebook,Google,TikTok,WhatsApp,Recomendación,Cliente recurrente,Otro"',"AH4:AH203")  # col 34
    add_dropdown(ws,'"Particular,Empresa,Recurrente,VIP"',"AI4:AI203")  # col 35

    # ── Formato condicional v1 (intacto) ──
    gf=PatternFill("solid",fgColor=VERDE_CLAR); rf=PatternFill("solid",fgColor=ROJO_CLAR)
    af=PatternFill("solid",fgColor=AMBAR_CLAR); bf=PatternFill("solid",fgColor=AZUL_CLAR)
    ws.conditional_formatting.add("Q4:Q203",CellIsRule(operator="equal",formula=['"Terminado"'],fill=gf,font=Font(color=VERDE_OSC,bold=True)))
    ws.conditional_formatting.add("Q4:Q203",CellIsRule(operator="equal",formula=['"Cancelado"'],fill=rf,font=Font(color=ROJO_OSC,bold=True)))
    ws.conditional_formatting.add("Q4:Q203",CellIsRule(operator="equal",formula=['"En proceso"'],fill=af,font=Font(color=AMBAR,bold=True)))
    ws.conditional_formatting.add("Q4:Q203",CellIsRule(operator="equal",formula=['"Confirmado"'],fill=bf,font=Font(color=AZUL_MED,bold=True)))

    # Formato condicional PRIORIDAD (col 23 = W)
    ws.conditional_formatting.add("W4:W203",CellIsRule(operator="equal",formula=['"Alta"'],fill=PatternFill("solid",fgColor=ROJO_CLAR),font=Font(color=ROJO_OSC,bold=True)))
    ws.conditional_formatting.add("W4:W203",CellIsRule(operator="equal",formula=['"Media"'],fill=PatternFill("solid",fgColor=AMBAR_CLAR),font=Font(color=AMBAR,bold=True)))
    ws.conditional_formatting.add("W4:W203",CellIsRule(operator="equal",formula=['"Baja"'],fill=PatternFill("solid",fgColor=VERDE_CLAR),font=Font(color=VERDE_OSC,bold=True)))

    return ws


# ── COSTOS (idéntico a v1) ────────────────────────────────────────────────────
def build_costos(wb):
    ws = wb.create_sheet("Costos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    merge_header(ws,1,1,20,"MUDANZAS METEPEC — COSTOS DEL SERVICIO",bg=VERDE_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,20,"Conectado por Folio con Agenda  |  Costo $7/km Pick Up · $9/km 1.5C · $10/km 2.5 · $12/km 3.5",bg=VERDE_MED,size=10)
    row_h(ws,2,22)
    headers=[("FOLIO",10),("CLIENTE",20),("UNIDAD",14),("KM REALES",10),("$/KM",8),
             ("COSTO UNIDAD",13),("CHOFER",12),("AYUDANTES",10),("MANIOBRAS",10),
             ("CASETAS",10),("COMIDAS",10),("HOTELES",10),("GASOLINA+",10),
             ("MATERIALES",11),("OTROS",10),("TOTAL GASTOS",13),("PRECIO COBRADO",14),
             ("UTILIDAD",12),("MARGEN %",10),("NOTAS",22)]
    for i,(h,w) in enumerate(headers,1):
        header_cell(ws,3,i,h,bg=VERDE_OSC); col_w(ws,i,w)
    row_h(ws,3,30)
    ws.cell(row=5,column=22,value="Tabla $/km").font=font(bold=True,color=VERDE_OSC)
    for i,(u,rv) in enumerate([("Pick Up",7),("1.5 Cerrada",9),("2.5",10),("3.5",12)],6):
        ws.cell(row=i,column=22,value=u)
        ws.cell(row=i,column=23,value=rv).number_format='"$"#,##0'
    for r in range(4,204):
        rb=BLANCO if r%2==0 else GRIS_CLAR
        c=ws.cell(row=r,column=1); c.fill=fill(AZUL_CLAR); c.font=font(bold=True,color=AZUL_OSC); c.alignment=align(h="center"); c.border=border_thin()
        c2=ws.cell(row=r,column=2,value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:E,5,0),"—"),"")')
        c2.fill=fill(rb); c2.border=border_thin(); c2.alignment=align(h="left")
        c3=ws.cell(row=r,column=3,value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:K,11,0),"—"),"")')
        c3.fill=fill(rb); c3.border=border_thin(); c3.alignment=align(h="center")
        c4=ws.cell(row=r,column=4); c4.fill=fill(rb); c4.border=border_thin(); c4.alignment=align(h="right"); c4.number_format="#,##0.0"
        km_f=(f'=IF(A{r}<>"",IF(C{r}="Pick Up",7,IF(C{r}="1.5 Cerrada",9,IF(C{r}="2.5",10,IF(C{r}="3.5",12,0)))),"")')
        c5=ws.cell(row=r,column=5,value=km_f); c5.fill=fill(AZUL_CLAR); c5.border=border_thin(); c5.alignment=align(h="right"); c5.number_format='"$"#,##0'
        c6=ws.cell(row=r,column=6,value=f'=IF(AND(D{r}<>"",E{r}<>""),D{r}*E{r},"")')
        c6.fill=fill(VERDE_CLAR); c6.font=font(bold=True,color=VERDE_OSC); c6.border=border_thin(); c6.number_format='"$"#,##0.00'; c6.alignment=align(h="right")
        for col in range(7,16):
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin(); c.number_format='"$"#,##0.00'; c.alignment=align(h="right")
        c16=ws.cell(row=r,column=16,value=f'=IF(A{r}<>"",IFERROR(F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r}+N{r}+O{r},0),"")')
        c16.fill=fill(ROJO_CLAR); c16.font=font(bold=True,color=ROJO_OSC); c16.border=border_thin(); c16.number_format='"$"#,##0.00'; c16.alignment=align(h="right")
        c17=ws.cell(row=r,column=17,value=f'=IF(A{r}<>"",IFERROR(VLOOKUP(A{r},Agenda!A:N,14,0),0),"")')
        c17.fill=fill(rb); c17.border=border_thin(); c17.number_format='"$"#,##0.00'; c17.alignment=align(h="right")
        c18=ws.cell(row=r,column=18,value=f'=IF(A{r}<>"",Q{r}-P{r},"")')
        c18.fill=fill(VERDE_CLAR); c18.font=font(bold=True,color=VERDE_OSC); c18.border=border_thin(); c18.number_format='"$"#,##0.00'; c18.alignment=align(h="right")
        c19=ws.cell(row=r,column=19,value=f'=IF(AND(A{r}<>"",Q{r}<>0),R{r}/Q{r}*100,"")')
        c19.fill=fill(rb); c19.border=border_thin(); c19.number_format='0.0"%"'; c19.alignment=align(h="right")
        c20=ws.cell(row=r,column=20); c20.fill=fill(rb); c20.border=border_thin(); c20.alignment=align(h="left",wrap=True)
        row_h(ws,r,18)
    ws.conditional_formatting.add("S4:S203",ColorScaleRule(start_type="num",start_value=0,start_color="FF0000",mid_type="num",mid_value=30,mid_color="FFFF00",end_type="num",end_value=60,end_color="00AA00"))
    return ws


# ── UNIDADES (v1 íntegra + expediente + indicadores + ocupación) ──────────────
def build_unidades(wb):
    ws = wb.create_sheet("Unidades")
    ws.sheet_view.showGridLines = False

    # == BLOQUE V1 INTACTO ==
    merge_header(ws,1,1,9,"MUDANZAS METEPEC — CONTROL POR UNIDAD",bg=AZUL_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,9,"Desempeño · Metas · Ocupación · Rentabilidad",bg=AZUL_MED,size=10)
    row_h(ws,2,22)
    merge_header(ws,3,1,9,"RESUMEN MENSUAL",bg=AZUL_OSC,size=11)
    row_h(ws,3,25)
    month_headers=["UNIDAD","SERVICIOS","VENTAS $","COSTOS $","UTILIDAD $","KM TOTAL","META MES $","CUMP. %","ESTADO"]
    widths=[18,12,14,14,14,12,14,12,14]
    for i,(h,w) in enumerate(zip(month_headers,widths),1):
        header_cell(ws,4,i,h,bg=AZUL_MED); col_w(ws,i,w)
    row_h(ws,4,28)
    units=[("Pick Up",24250),("1.5 Cerrada",28450),("2.5",31250),("3.5",15250)]
    for i,(uname,meta_mes) in enumerate(units,5):
        rb=BLANCO if i%2==0 else GRIS_CLAR; u=f'"{uname}"'
        c=ws.cell(row=i,column=1,value=uname); c.fill=fill(AZUL_CLAR); c.font=font(bold=True,color=AZUL_OSC); c.border=border_thin(); c.alignment=align(h="center")
        sv=ws.cell(row=i,column=2,value=f'=COUNTIF(Agenda!K:K,{u})'); sv.fill=fill(rb); sv.border=border_thin(); sv.alignment=align(h="center"); sv.font=font(bold=True)
        vt=ws.cell(row=i,column=3,value=f'=SUMIF(Agenda!K:K,{u},Agenda!N:N)'); vt.fill=fill(rb); vt.border=border_thin(); vt.number_format='"$"#,##0.00'; vt.alignment=align(h="right"); vt.font=font(bold=True,color=VERDE_OSC)
        ct=ws.cell(row=i,column=4,value=f'=SUMIF(Costos!C:C,{u},Costos!P:P)'); ct.fill=fill(rb); ct.border=border_thin(); ct.number_format='"$"#,##0.00'; ct.alignment=align(h="right"); ct.font=font(color=ROJO_OSC)
        ut=ws.cell(row=i,column=5,value=f'=C{i}-D{i}'); ut.fill=fill(VERDE_CLAR); ut.border=border_thin(); ut.number_format='"$"#,##0.00'; ut.alignment=align(h="right"); ut.font=font(bold=True,color=VERDE_OSC)
        km=ws.cell(row=i,column=6,value=f'=SUMIF(Costos!C:C,{u},Costos!D:D)'); km.fill=fill(rb); km.border=border_thin(); km.number_format='#,##0.0" km"'; km.alignment=align(h="right")
        mt=ws.cell(row=i,column=7,value=meta_mes); mt.fill=fill(AMBAR_CLAR); mt.border=border_thin(); mt.number_format='"$"#,##0.00'; mt.alignment=align(h="right"); mt.font=font(bold=True,color=AMBAR)
        cp=ws.cell(row=i,column=8,value=f'=IF(G{i}>0,C{i}/G{i}*100,0)'); cp.fill=fill(rb); cp.border=border_thin(); cp.number_format='0.0"%"'; cp.alignment=align(h="right"); cp.font=font(bold=True)
        est=ws.cell(row=i,column=9,value=f'=IF(H{i}>=100,"✅ META CUMPLIDA",IF(H{i}>=70,"⚠️ En camino","🔴 Atención"))'); est.fill=fill(rb); est.border=border_thin(); est.alignment=align(h="center")
        row_h(ws,i,22)
    r_tot=9
    merge_header(ws,r_tot,1,1,"TOTAL FLOTA",bg=GRIS_OSC,fg=BLANCO,size=10)
    ws.cell(row=r_tot,column=1).alignment=align(h="center")
    for col,formula in [(2,"=SUM(B5:B8)"),(3,"=SUM(C5:C8)"),(4,"=SUM(D5:D8)"),(5,"=SUM(E5:E8)"),(6,"=SUM(F5:F8)"),(7,"=SUM(G5:G8)")]:
        c=ws.cell(row=r_tot,column=col,value=formula); c.fill=fill(GRIS_OSC); c.font=font(bold=True,color=BLANCO); c.border=border_thin(); c.alignment=align(h="right" if col>2 else "center")
        if col in [3,4,5,7]: c.number_format='"$"#,##0.00'
        elif col==6: c.number_format='#,##0.0" km"'
    row_h(ws,r_tot,22)
    r_sem=11
    merge_header(ws,r_sem,1,9,"RESUMEN SEMANAL",bg=AZUL_OSC,size=11)
    row_h(ws,r_sem,25)
    week_headers=["UNIDAD","SERVICIOS SEMANA","VENTAS SEMANA $","META SEMANA $","CUMP. SEMANAL %","FALTA $","PROM. x SERVICIO $","MEJOR SEMANA","NOTA"]
    for i,h in enumerate(week_headers,1): header_cell(ws,r_sem+1,i,h,bg=VERDE_OSC)
    row_h(ws,r_sem+1,28)
    units_week=[("Pick Up",5600),("1.5 Cerrada",6570),("2.5",7220),("3.5",3520)]
    for idx,(uname,meta_sem) in enumerate(units_week):
        r=r_sem+2+idx; rb=BLANCO if idx%2==0 else GRIS_CLAR; u=f'"{uname}"'
        c1=ws.cell(row=r,column=1,value=uname); c1.fill=fill(AZUL_CLAR); c1.font=font(bold=True,color=AZUL_OSC); c1.border=border_thin(); c1.alignment=align(h="center")
        sv=ws.cell(row=r,column=2,value=f'=COUNTIFS(Agenda!K:K,{u},Agenda!C:C,">="&TODAY()-7,Agenda!C:C,"<="&TODAY())'); sv.fill=fill(rb); sv.border=border_thin(); sv.alignment=align(h="center"); sv.font=font(bold=True)
        vt=ws.cell(row=r,column=3,value=f'=SUMIFS(Agenda!N:N,Agenda!K:K,{u},Agenda!C:C,">="&TODAY()-7,Agenda!C:C,"<="&TODAY())'); vt.fill=fill(rb); vt.border=border_thin(); vt.number_format='"$"#,##0.00'; vt.alignment=align(h="right"); vt.font=font(bold=True,color=VERDE_OSC)
        mt=ws.cell(row=r,column=4,value=meta_sem); mt.fill=fill(AMBAR_CLAR); mt.border=border_thin(); mt.number_format='"$"#,##0.00'; mt.alignment=align(h="right"); mt.font=font(bold=True,color=AMBAR)
        cp=ws.cell(row=r,column=5,value=f'=IF(D{r}>0,C{r}/D{r}*100,0)'); cp.fill=fill(rb); cp.border=border_thin(); cp.number_format='0.0"%"'; cp.alignment=align(h="right"); cp.font=font(bold=True)
        flt=ws.cell(row=r,column=6,value=f'=MAX(0,D{r}-C{r})'); flt.fill=fill(ROJO_CLAR); flt.border=border_thin(); flt.number_format='"$"#,##0.00'; flt.alignment=align(h="right"); flt.font=font(color=ROJO_OSC)
        prm=ws.cell(row=r,column=7,value=f'=IF(B{r}>0,C{r}/B{r},0)'); prm.fill=fill(rb); prm.border=border_thin(); prm.number_format='"$"#,##0.00'; prm.alignment=align(h="right")
        ws.cell(row=r,column=8,value="").border=border_thin()
        ws.cell(row=r,column=9,value="").border=border_thin()
        row_h(ws,r,22)

    # == NUEVO: EXPEDIENTE TÉCNICO POR UNIDAD (rows 18+) ==
    R = 18
    merge_header(ws, R, 1, 6, "EXPEDIENTE TÉCNICO DE UNIDADES", bg=AZUL_OSC, size=11)
    row_h(ws, R, 28)
    R += 1
    merge_header(ws, R, 1, 6, "Datos técnicos · Documentación · Kilometraje · Alertas", bg=AZUL_MED, size=10)
    row_h(ws, R, 22)
    R += 1

    # Encabezados: col1=campo, cols 2-5 = una por unidad
    unit_cols = ["Pick Up", "1.5 Cerrada", "2.5", "3.5"]
    header_cell(ws, R, 1, "CAMPO", bg=GRIS_OSC)
    for i, uname in enumerate(unit_cols, 2):
        header_cell(ws, R, i, uname, bg=AZUL_MED)
    header_cell(ws, R, 6, "ALERTA", bg=ROJO_OSC)
    row_h(ws, R, 26)
    R += 1

    col_w(ws, 1, 28); col_w(ws, 6, 22)

    exp_rows = [
        ("Fecha de compra",         None, "DD/MM/YYYY"),
        ("Valor de compra $",       None, '"$"#,##0.00'),
        ("Kilometraje inicial",     None, '#,##0" km"'),
        ("Kilometraje actual",      None, '#,##0" km"'),
        ("KM recorridos",           "=IF(AND(B{km_ini}<>\"\",B{km_act}<>\"\"),B{km_act}-B{km_ini},\"\")", '#,##0" km"'),
        ("Vida útil objetivo",      300000, '#,##0" km"'),
        ("% vida útil consumida",   "=IF(B{km_rec}<>\"\",IFERROR(B{km_rec}/B{v_util}*100,0),\"\")", '0.0"%"'),
        ("KM restantes",            "=IF(B{km_rec}<>\"\",IFERROR(B{v_util}-B{km_rec},0),\"\")", '#,##0" km"'),
        ("Fecha próximo servicio",  None, "DD/MM/YYYY"),
        ("KM para próx. servicio",  None, '#,##0" km"'),
        ("Próx. cambio de llantas", None, "DD/MM/YYYY"),
        ("Vencimiento seguro",      None, "DD/MM/YYYY"),
        ("Próxima verificación",    None, "DD/MM/YYYY"),
        ("Vencimiento tenencia",    None, "DD/MM/YYYY"),
        ("GPS instalado",           None, None),
        ("Observaciones unidad",    None, None),
    ]

    # Rows para referencias de fórmulas (posición relativa)
    km_ini_r = R + 2   # fila de km inicial
    km_act_r = R + 3   # fila de km actual
    km_rec_r = R + 4   # fila de km recorridos
    v_util_r = R + 5   # fila vida útil

    for fi, (label, default, nfmt) in enumerate(exp_rows):
        row_r = R + fi
        rb = BLANCO if fi % 2 == 0 else GRIS_CLAR

        c = ws.cell(row=row_r, column=1, value=label)
        c.fill = fill(GRIS_CLAR); c.font = font(bold=True, color=GRIS_OSC)
        c.border = border_thin(); c.alignment = align(h="left")

        for col_i in range(2, 6):  # cols 2-5 (una por unidad)
            val = default
            # fórmulas calculadas automáticamente
            if label == "KM recorridos":
                val = f'=IF(AND({get_column_letter(col_i)}{km_ini_r}<>"",{get_column_letter(col_i)}{km_act_r}<>""),{get_column_letter(col_i)}{km_act_r}-{get_column_letter(col_i)}{km_ini_r},"")'
            elif label == "% vida útil consumida":
                val = f'=IF({get_column_letter(col_i)}{km_rec_r}<>"",IFERROR({get_column_letter(col_i)}{km_rec_r}/{get_column_letter(col_i)}{v_util_r}*100,0),"")'
            elif label == "KM restantes":
                val = f'=IF({get_column_letter(col_i)}{km_rec_r}<>"",IFERROR({get_column_letter(col_i)}{v_util_r}-{get_column_letter(col_i)}{km_rec_r},0),"")'

            c = ws.cell(row=row_r, column=col_i, value=val)
            c.fill = fill(rb); c.border = border_thin()
            if nfmt: c.number_format = nfmt
            c.alignment = align(h="right" if nfmt and "$" in nfmt else "center")
            if label == "Valor de compra $": c.font = font(color=ROJO_OSC)
            if label in ["KM recorridos","% vida útil consumida","KM restantes"]:
                c.fill = fill(AZUL_CLAR); c.font = font(color=AZUL_OSC)

        # Alerta col 6
        if label == "% vida útil consumida":
            for col_i in range(2, 6):
                cl = get_column_letter(col_i)
                al = ws.cell(row=row_r, column=6,
                    value=f'=IF(MAX(B{row_r}:E{row_r})>=90,"🔴 CAMBIAR",IF(MAX(B{row_r}:E{row_r})>=70,"⚠️ Revisar","✅ OK"))')
                al.fill = fill(AMBAR_CLAR); al.font = font(bold=True, color=AMBAR)
                al.border = border_thin(); al.alignment = align(h="center")
                break
        elif label in ["Vencimiento seguro","Próxima verificación","Vencimiento tenencia"]:
            al = ws.cell(row=row_r, column=6,
                value=f'=IF(MIN(IF(B{row_r}:E{row_r}<>"",B{row_r}:E{row_r},99999))-TODAY()<=30,"🔴 Vence pronto","✅ OK")')
            al.fill = fill(rb); al.font = font(color=GRIS_MED, italic=True)
            al.border = border_thin(); al.alignment = align(h="center")
        else:
            ws.cell(row=row_r, column=6).border = border_thin()
            ws.cell(row=row_r, column=6).fill = fill(rb)

        row_h(ws, row_r, 20)

    R += len(exp_rows)

    # == NUEVO: INDICADORES POR UNIDAD ==
    R += 1
    merge_header(ws, R, 1, 6, "INDICADORES POR UNIDAD (Acumulado)", bg=VERDE_OSC, size=11)
    row_h(ws, R, 28)
    R += 1
    header_cell(ws, R, 1, "INDICADOR", bg=VERDE_OSC)
    for i, uname in enumerate(unit_cols, 2):
        header_cell(ws, R, i, uname, bg=VERDE_MED)
    ws.cell(row=R, column=6).border = border_thin()
    row_h(ws, R, 26)
    R += 1

    # fila de km recorridos en el expediente (para calcular $/km)
    km_rec_exp_row = 18 + 3 + 4 + 4  # row 18 (header) + 1 sub-header + 1 col-header = R_exp_start+4 = km_rec row
    # Recalculo: expediente starts at row 18(header),19(sub),20(col-heads),21+(data)
    # km_rec es el 5to campo (index 4) → fila 21+4 = 25
    km_rec_exp = 25  # fila donde está KM recorridos en expediente

    ind_rows = [
        ("Ventas acumuladas",        [f'=SUMIF(Agenda!K:K,"{u}",Agenda!N:N)' for u in unit_cols], '"$"#,##0.00', VERDE_CLAR, VERDE_OSC),
        ("Costos acumulados",        [f'=SUMIF(Costos!C:C,"{u}",Costos!P:P)' for u in unit_cols], '"$"#,##0.00', ROJO_CLAR, ROJO_OSC),
        ("Utilidad acumulada",       [f'=SUMIF(Agenda!K:K,"{u}",Agenda!N:N)-SUMIF(Costos!C:C,"{u}",Costos!P:P)' for u in unit_cols], '"$"#,##0.00', VERDE_CLAR, VERDE_OSC),
        ("Servicios realizados",     [f'=COUNTIF(Agenda!K:K,"{u}")' for u in unit_cols], "#,##0", AZUL_CLAR, AZUL_OSC),
        ("Promedio por servicio $",  None, '"$"#,##0.00', AZUL_CLAR, AZUL_OSC),
        ("Ingreso por km $",         None, '"$"#,##0.00', GRIS_CLAR, GRIS_OSC),
        ("Costo por km real $",      None, '"$"#,##0.00', ROJO_CLAR, ROJO_OSC),
        ("Utilidad por km $",        None, '"$"#,##0.00', VERDE_CLAR, VERDE_OSC),
        ("Meta mensual $",           [24250, 28450, 31250, 15250], '"$"#,##0.00', AMBAR_CLAR, AMBAR),
        ("Meta semanal $",           [5600,  6570,  7220,  3520],  '"$"#,##0.00', AMBAR_CLAR, AMBAR),
        ("% Cumplimiento mensual",   None, '0.0"%"', AZUL_CLAR, AZUL_OSC),
    ]

    ind_start = R
    for fi, (label, formulas, nfmt, bg_c, fg_c) in enumerate(ind_rows):
        row_r = R + fi
        rb = BLANCO if fi % 2 == 0 else GRIS_CLAR

        c = ws.cell(row=row_r, column=1, value=label)
        c.fill = fill(GRIS_CLAR); c.font = font(bold=True, color=GRIS_OSC)
        c.border = border_thin(); c.alignment = align(h="left")

        for col_i in range(2, 6):
            u = unit_cols[col_i - 2]
            cl = get_column_letter(col_i)

            if formulas is not None:
                val = formulas[col_i - 2]
            elif label == "Promedio por servicio $":
                sv_row = ind_start + 3   # servicios realizados row
                vt_row = ind_start       # ventas acumuladas row
                val = f'=IF({cl}{sv_row}>0,{cl}{vt_row}/{cl}{sv_row},0)'
            elif label == "Ingreso por km $":
                vt_row = ind_start
                val = f'=IF({cl}{km_rec_exp}<>"",IFERROR({cl}{vt_row}/{cl}{km_rec_exp},0),0)'
            elif label == "Costo por km real $":
                ct_row = ind_start + 1
                val = f'=IF({cl}{km_rec_exp}<>"",IFERROR({cl}{ct_row}/{cl}{km_rec_exp},0),0)'
            elif label == "Utilidad por km $":
                ut_row = ind_start + 2
                val = f'=IF({cl}{km_rec_exp}<>"",IFERROR({cl}{ut_row}/{cl}{km_rec_exp},0),0)'
            elif label == "% Cumplimiento mensual":
                vt_row = ind_start
                meta_row = ind_start + 8
                val = f'=IF({cl}{meta_row}>0,{cl}{vt_row}/{cl}{meta_row}*100,0)'
            else:
                val = None

            c = ws.cell(row=row_r, column=col_i, value=val)
            c.fill = fill(bg_c if formulas is not None or val else rb)
            c.font = font(bold=(label in ["Ventas acumuladas","Utilidad acumulada"]), color=fg_c)
            c.border = border_thin()
            c.number_format = nfmt
            c.alignment = align(h="right")

        ws.cell(row=row_r, column=6).border = border_thin()
        ws.cell(row=row_r, column=6).fill = fill(rb)
        row_h(ws, row_r, 20)

    R += len(ind_rows)

    # == NUEVO: OCUPACIÓN ==
    R += 1
    merge_header(ws, R, 1, 6, "OCUPACIÓN DE UNIDADES", bg=GRIS_OSC, size=11)
    row_h(ws, R, 28)
    R += 1
    header_cell(ws, R, 1, "INDICADOR", bg=GRIS_OSC)
    for i, uname in enumerate(unit_cols, 2):
        header_cell(ws, R, i, uname, bg=GRIS_MED)
    ws.cell(row=R, column=6).border = border_thin()
    row_h(ws, R, 26)
    R += 1

    ocup_rows = [
        ("Días trabajados (mes)",
         [f'=COUNTIFS(Agenda!K:K,"{u}",Agenda!Q:Q,"Terminado",Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))' for u in unit_cols],
         "#,##0", AZUL_CLAR, AZUL_OSC),
        ("Días del mes",
         ["=DAY(EOMONTH(TODAY(),0))"] * 4,
         "#,##0", GRIS_CLAR, GRIS_OSC),
        ("% Ocupación mensual",
         None, '0.0"%"', VERDE_CLAR, VERDE_OSC),
        ("Servicios esta semana",
         [f'=COUNTIFS(Agenda!K:K,"{u}",Agenda!C:C,">="&TODAY()-7,Agenda!C:C,"<="&TODAY())' for u in unit_cols],
         "#,##0", AZUL_CLAR, AZUL_OSC),
        ("Servicios terminados total",
         [f'=COUNTIFS(Agenda!K:K,"{u}",Agenda!Q:Q,"Terminado")' for u in unit_cols],
         "#,##0", VERDE_CLAR, VERDE_OSC),
        ("Servicios cancelados total",
         [f'=COUNTIFS(Agenda!K:K,"{u}",Agenda!Q:Q,"Cancelado")' for u in unit_cols],
         "#,##0", ROJO_CLAR, ROJO_OSC),
        ("Prom. servicios/semana",
         None, '0.0" sv/sem"', GRIS_CLAR, GRIS_OSC),
    ]

    dias_row = R
    for fi, (label, formulas, nfmt, bg_c, fg_c) in enumerate(ocup_rows):
        row_r = R + fi
        rb = BLANCO if fi % 2 == 0 else GRIS_CLAR

        c = ws.cell(row=row_r, column=1, value=label)
        c.fill = fill(GRIS_CLAR); c.font = font(bold=True, color=GRIS_OSC)
        c.border = border_thin(); c.alignment = align(h="left")

        for col_i in range(2, 6):
            cl = get_column_letter(col_i)
            if formulas is not None:
                val = formulas[col_i - 2]
            elif label == "% Ocupación mensual":
                val = f'=IF({cl}{dias_row+1}>0,{cl}{dias_row}/{cl}{dias_row+1}*100,0)'
            elif label == "Prom. servicios/semana":
                tot_row = dias_row + 4
                val = f'=IFERROR({cl}{tot_row}/IFERROR(DATEDIF(MINIFS(Agenda!C:C,Agenda!K:K,"{unit_cols[col_i-2]}"),TODAY(),"W"),1),0)'
            else:
                val = None

            c = ws.cell(row=row_r, column=col_i, value=val)
            c.fill = fill(bg_c)
            c.font = font(bold=(label == "% Ocupación mensual"), color=fg_c)
            c.border = border_thin(); c.number_format = nfmt
            c.alignment = align(h="right")

        ws.cell(row=row_r, column=6).border = border_thin()
        ws.cell(row=row_r, column=6).fill = fill(rb)
        row_h(ws, row_r, 20)

    return ws


# ── FLUJO (idéntico a v1 - ya tenía Folio y Unidad) ─────────────────────────
def build_flujo(wb):
    ws = wb.create_sheet("Flujo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    merge_header(ws,1,1,9,"MUDANZAS METEPEC — FLUJO DE EFECTIVO",bg=VERDE_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,9,"Cada movimiento de dinero: entradas · salidas · saldo acumulado",bg=VERDE_MED,size=10)
    row_h(ws,2,22)
    headers=[("FECHA",12),("CONCEPTO",35),("ENTRADA $",14),("SALIDA $",14),
             ("CATEGORÍA",22),("UNIDAD",16),("FOLIO",12),("MÉTODO PAGO",14),("SALDO ACUM. $",16)]
    for i,(h,w) in enumerate(headers,1):
        header_cell(ws,3,i,h,bg=VERDE_OSC); col_w(ws,i,w)
    row_h(ws,3,30)
    c_si=ws.cell(row=4,column=2,value="SALDO INICIAL"); c_si.fill=fill(AZUL_CLAR); c_si.font=font(bold=True,color=AZUL_OSC); c_si.border=border_thin(); c_si.alignment=align(h="left")
    for col in [1,3,4,5,6,7,8]:
        c=ws.cell(row=4,column=col); c.fill=fill(AZUL_CLAR); c.border=border_thin()
        if col in [3,4]: c.number_format='"$"#,##0.00'; c.alignment=align(h="right")
    c0=ws.cell(row=4,column=9,value='=IF(C4<>"",C4,0)'); c0.fill=fill(VERDE_CLAR); c0.font=font(bold=True,color=VERDE_OSC); c0.border=border_thin(); c0.number_format='"$"#,##0.00'; c0.alignment=align(h="right")
    row_h(ws,4,22)
    for r in range(5,505):
        rb=BLANCO if r%2==0 else GRIS_CLAR
        cf=ws.cell(row=r,column=1); cf.fill=fill(rb); cf.border=border_thin(); cf.number_format="DD/MM/YYYY"; cf.alignment=align(h="center")
        cc=ws.cell(row=r,column=2); cc.fill=fill(rb); cc.border=border_thin(); cc.alignment=align(h="left")
        ce=ws.cell(row=r,column=3); ce.fill=fill(VERDE_CLAR); ce.border=border_thin(); ce.number_format='"$"#,##0.00'; ce.alignment=align(h="right")
        cs=ws.cell(row=r,column=4); cs.fill=fill(ROJO_CLAR); cs.border=border_thin(); cs.number_format='"$"#,##0.00'; cs.alignment=align(h="right")
        for col in [5,6,7,8]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin(); c.alignment=align(h="center")
        ca=ws.cell(row=r,column=9,value=f'=IF(OR(C{r}<>"",D{r}<>""),I{r-1}+IF(C{r}<>"",C{r},0)-IF(D{r}<>"",D{r},0),"")')
        ca.fill=fill(VERDE_CLAR); ca.font=font(bold=True,color=VERDE_OSC); ca.border=border_thin(); ca.number_format='"$"#,##0.00'; ca.alignment=align(h="right")
        row_h(ws,r,18)
    add_dropdown(ws,'"Servicio,Gasto operativo,Mensualidad,Publicidad,Administración,Deuda,Nómina,Mantenimiento"',"E4:E504")
    add_dropdown(ws,'"Pick Up,1.5 Cerrada,2.5,3.5,General"',"F4:F504")
    add_dropdown(ws,'"Efectivo,Transferencia,Tarjeta,Mixto"',"H4:H504")
    col_s=11
    ws.column_dimensions[get_column_letter(col_s)].width=24
    ws.column_dimensions[get_column_letter(col_s+1)].width=16
    items=[("RESUMEN FLUJO",None,AZUL_OSC,BLANCO,True),("Total Entradas","=IFERROR(SUM(C4:C504),0)",VERDE_CLAR,VERDE_OSC,False),("Total Salidas","=IFERROR(SUM(D4:D504),0)",ROJO_CLAR,ROJO_OSC,False),("Saldo Neto","=K12-K13",AZUL_CLAR,AZUL_OSC,True),("",None,BLANCO,NEGRO,False),("Por categoría:",None,GRIS_CLAR,GRIS_OSC,True),("Servicios",'=SUMIF(E4:E504,"Servicio",C4:C504)',VERDE_CLAR,VERDE_OSC,False),("G. Operativo",'=SUMIF(E4:E504,"Gasto operativo",D4:D504)',ROJO_CLAR,ROJO_OSC,False),("Nómina",'=SUMIF(E4:E504,"Nómina",D4:D504)',ROJO_CLAR,ROJO_OSC,False),("Publicidad",'=SUMIF(E4:E504,"Publicidad",D4:D504)',AMBAR_CLAR,AMBAR,False),("Mantenimiento",'=SUMIF(E4:E504,"Mantenimiento",D4:D504)',ROJO_CLAR,ROJO_OSC,False)]
    for idx,(label,formula,bg_c,fg_c,bold_) in enumerate(items,11):
        rr=idx; c=ws.cell(row=rr,column=col_s,value=label); c.fill=fill(bg_c); c.font=font(bold=bold_,color=fg_c); c.border=border_thin(); c.alignment=align(h="left"); row_h(ws,rr,22)
        if formula:
            cv=ws.cell(row=rr,column=col_s+1,value=formula); cv.fill=fill(bg_c); cv.font=font(bold=bold_,color=fg_c); cv.border=border_thin(); cv.number_format='"$"#,##0.00'; cv.alignment=align(h="right")
    return ws


# ── GASTOS FIJOS (idéntico a v1) ─────────────────────────────────────────────
def build_gastos_fijos(wb):
    ws = wb.create_sheet("Gastos Fijos")
    ws.sheet_view.showGridLines = False
    merge_header(ws,1,1,5,"MUDANZAS METEPEC — GASTOS FIJOS",bg=ROJO_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,5,"Gastos fijos mensuales · Punto de equilibrio · Avance de cobertura",bg="C0392B",size=10)
    row_h(ws,2,22)
    col_w(ws,1,30); col_w(ws,2,16); col_w(ws,3,16); col_w(ws,4,16); col_w(ws,5,22)
    for i,h in enumerate(["CONCEPTO","MONTO MES $","MONTO SEM $","% DEL TOTAL","CATEGORÍA"],1):
        header_cell(ws,3,i,h,bg=ROJO_OSC)
    row_h(ws,3,28)
    gastos=[("Camioneta 2.5",16000,"Unidad"),("Camioneta 1.5 C",13200,"Unidad"),("Camioneta Pick Up",9000,"Unidad"),("Pensión (estacionam)",20000,"Operativo"),("Anuncios / Publicidad",6000,"Marketing"),("Suscripciones",5000,"Admin"),("Administración",30000,"Admin")]
    for idx,(concepto,monto,cat) in enumerate(gastos,4):
        rb=BLANCO if idx%2==0 else GRIS_CLAR; r=idx
        c1=ws.cell(row=r,column=1,value=concepto); c1.fill=fill(rb); c1.font=font(color=GRIS_OSC); c1.border=border_thin(); c1.alignment=align(h="left")
        c2=ws.cell(row=r,column=2,value=monto); c2.fill=fill(rb); c2.border=border_thin(); c2.number_format='"$"#,##0.00'; c2.alignment=align(h="right"); c2.font=font(color=ROJO_OSC)
        c3=ws.cell(row=r,column=3,value=f'=B{r}/4.33'); c3.fill=fill(rb); c3.border=border_thin(); c3.number_format='"$"#,##0.00'; c3.alignment=align(h="right")
        c4=ws.cell(row=r,column=4,value=f'=B{r}/$B$11*100'); c4.fill=fill(rb); c4.border=border_thin(); c4.number_format='0.0"%"'; c4.alignment=align(h="right")
        c5=ws.cell(row=r,column=5,value=cat); c5.fill=fill(rb); c5.border=border_thin(); c5.alignment=align(h="center")
        row_h(ws,r,22)
    r_tot=11
    for col,val,nfmt in [(1,"TOTAL MENSUAL",None),(2,"=SUM(B4:B10)",'"$"#,##0.00'),(3,"=B11/4.33",'"$"#,##0.00'),(4,"100%",None)]:
        c=ws.cell(row=r_tot,column=col,value=val); c.fill=fill(ROJO_OSC); c.font=font(bold=True,color=BLANCO,size=11 if col<=2 else 10); c.border=border_thin()
        if nfmt: c.number_format=nfmt
        c.alignment=align(h="right" if col>1 else "left")
    row_h(ws,r_tot,26)
    r_panel=13
    merge_header(ws,r_panel,1,5,"AVANCE DE COBERTURA DE GASTOS FIJOS",bg=AZUL_OSC,size=11)
    row_h(ws,r_panel,28)
    cov=[("Ventas del mes (total flota)","=IFERROR(SUM(Agenda!N:N),0)",VERDE_CLAR,VERDE_OSC),("Gastos fijos totales","=B11",ROJO_CLAR,ROJO_OSC),("Gastos fijos CUBIERTOS","=MIN(B14,B15)",VERDE_CLAR,VERDE_OSC),("FALTA para cubrir gastos fijos","=MAX(0,B15-B14)",AMBAR_CLAR,AMBAR),("% Cobertura alcanzado","=IF(B15>0,B14/B15*100,0)",AZUL_CLAR,AZUL_OSC),("Ventas necesarias p/equilibrio","=B15",GRIS_CLAR,GRIS_MED),("Margen sobre gastos fijos","=MAX(0,B14-B15)",VERDE_CLAR,VERDE_OSC)]
    for idx,(label,formula,bg_c,fg_c) in enumerate(cov,r_panel+1):
        r=idx
        c1=ws.cell(row=r,column=1,value=label); c1.fill=fill(bg_c); c1.font=font(color=fg_c); c1.border=border_thin(); c1.alignment=align(h="left")
        c2=ws.cell(row=r,column=2,value=formula); c2.fill=fill(bg_c); c2.font=font(bold=True,color=fg_c); c2.border=border_thin(); c2.number_format=('"$"#,##0.00' if "%" not in label else '0.0"%"'); c2.alignment=align(h="right")
        for col in [3,4,5]: ws.cell(row=r,column=col).fill=fill(bg_c); ws.cell(row=r,column=col).border=border_thin()
        row_h(ws,r,22)
    r_dist=r_panel+len(cov)+2
    merge_header(ws,r_dist,1,5,"DISTRIBUCIÓN META POR UNIDAD (vs. Gastos Fijos)",bg=AZUL_MED,size=11)
    row_h(ws,r_dist,25)
    dist_headers=["UNIDAD","META MES $","META SEM $","% META/GASTO","VENTAS REALES $"]
    for i,h in enumerate(dist_headers,1): header_cell(ws,r_dist+1,i,h,bg=AZUL_MED)
    row_h(ws,r_dist+1,26)
    dist_data=[("Pick Up",24250,5600),("1.5 Cerrada",28450,6570),("2.5",31250,7220),("3.5",15250,3520)]
    for idx,(u,meta_m,meta_s) in enumerate(dist_data,r_dist+2):
        rb=BLANCO if idx%2==0 else GRIS_CLAR; uq=f'"{u}"'
        ws.cell(row=idx,column=1,value=u).fill=fill(AZUL_CLAR); ws.cell(row=idx,column=1).font=font(bold=True,color=AZUL_OSC); ws.cell(row=idx,column=1).border=border_thin(); ws.cell(row=idx,column=1).alignment=align(h="center")
        ws.cell(row=idx,column=2,value=meta_m).fill=fill(rb); ws.cell(row=idx,column=2).number_format='"$"#,##0.00'; ws.cell(row=idx,column=2).border=border_thin(); ws.cell(row=idx,column=2).alignment=align(h="right")
        ws.cell(row=idx,column=3,value=meta_s).fill=fill(rb); ws.cell(row=idx,column=3).number_format='"$"#,##0.00'; ws.cell(row=idx,column=3).border=border_thin(); ws.cell(row=idx,column=3).alignment=align(h="right")
        ws.cell(row=idx,column=4,value=f'=B{idx}/$B$11*100').fill=fill(rb); ws.cell(row=idx,column=4).number_format='0.0"%"'; ws.cell(row=idx,column=4).border=border_thin(); ws.cell(row=idx,column=4).alignment=align(h="right")
        ws.cell(row=idx,column=5,value=f'=SUMIF(Agenda!K:K,{uq},Agenda!N:N)').fill=fill(VERDE_CLAR); ws.cell(row=idx,column=5).font=font(bold=True,color=VERDE_OSC); ws.cell(row=idx,column=5).number_format='"$"#,##0.00'; ws.cell(row=idx,column=5).border=border_thin(); ws.cell(row=idx,column=5).alignment=align(h="right")
        row_h(ws,idx,22)
    return ws


# ── CLIENTES (idéntico a v1) ──────────────────────────────────────────────────
def build_clientes(wb):
    ws = wb.create_sheet("Clientes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    merge_header(ws,1,1,16,"MUDANZAS METEPEC — CLIENTES Y SEGUIMIENTO",bg=AZUL_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,16,"CRM básico: contacto → cotización → cierre → reseña → recurrencia",bg=AZUL_MED,size=10)
    row_h(ws,2,22)
    headers=[("NOMBRE",22),("TELÉFONO",14),("F.CONTACTO",12),("FUENTE",16),("SERVICIO SOL.",22),("COTIZACIÓN $",13),("CERRÓ",9),("MOTIVO NO CERRÓ",22),("F.SEGUIMIENTO",13),("RESEÑA SOL.",11),("RESEÑA REC.",11),("RECURRENTE",11),("VALOR CLIENTE $",14),("SERVICIOS TOTAL",12),("ÚLTIMO SERVICIO",13),("OBSERVACIONES",30)]
    for i,(h,w) in enumerate(headers,1): header_cell(ws,3,i,h); col_w(ws,i,w)
    row_h(ws,3,30)
    for r in range(4,304):
        rb=BLANCO if r%2==0 else GRIS_CLAR
        for col in [1,2,5,8,16]: c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin(); c.alignment=align(h="left",wrap=(col in [8,16]))
        for col in [3,9,15]: c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin(); c.number_format="DD/MM/YYYY"; c.alignment=align(h="center")
        for col in [4,6,7,10,11,12]: c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin(); c.alignment=align(h="center")
        ws.cell(row=r,column=6).number_format='"$"#,##0.00'; ws.cell(row=r,column=6).alignment=align(h="right")
        vc=ws.cell(row=r,column=13,value=f'=IF(A{r}<>"",IFERROR(SUMIF(Agenda!E:E,A{r},Agenda!N:N),0),"")'); vc.fill=fill(VERDE_CLAR); vc.font=font(bold=True,color=VERDE_OSC); vc.border=border_thin(); vc.number_format='"$"#,##0.00'; vc.alignment=align(h="right")
        st=ws.cell(row=r,column=14,value=f'=IF(A{r}<>"",IFERROR(COUNTIF(Agenda!E:E,A{r}),0),"")'); st.fill=fill(AZUL_CLAR); st.font=font(bold=True,color=AZUL_OSC); st.border=border_thin(); st.alignment=align(h="center")
        us=ws.cell(row=r,column=15,value=f'=IF(A{r}<>"",IFERROR(MAXIFS(Agenda!C:C,Agenda!E:E,A{r}),""),"")'); us.fill=fill(rb); us.border=border_thin(); us.number_format="DD/MM/YYYY"; us.alignment=align(h="center")
        row_h(ws,r,18)
    add_dropdown(ws,'"Saleads,Facebook,WhatsApp,Recomendación,Google,Otro"',"D4:D303")
    add_dropdown(ws,'"Sí,No"',"G4:G303"); add_dropdown(ws,'"Sí,No"',"J4:J303")
    add_dropdown(ws,'"Sí,No"',"K4:K303"); add_dropdown(ws,'"Sí,No"',"L4:L303")
    ws.conditional_formatting.add("G4:G303",CellIsRule(operator="equal",formula=['"Sí"'],fill=PatternFill("solid",fgColor=VERDE_CLAR),font=Font(color=VERDE_OSC,bold=True)))
    ws.conditional_formatting.add("G4:G303",CellIsRule(operator="equal",formula=['"No"'],fill=PatternFill("solid",fgColor=ROJO_CLAR),font=Font(color=ROJO_OSC,bold=True)))
    col_m=18
    ws.column_dimensions[get_column_letter(col_m)].width=28
    ws.column_dimensions[get_column_letter(col_m+1)].width=16
    crm=[("MÉTRICAS CRM",None,AZUL_OSC,BLANCO),("Total contactos","=COUNTA(A4:A303)",AZUL_CLAR,AZUL_OSC),("Total cerrados",'=COUNTIF(G4:G303,"Sí")',VERDE_CLAR,VERDE_OSC),("No cerrados",'=COUNTIF(G4:G303,"No")',ROJO_CLAR,ROJO_OSC),("Tasa de cierre %",'=IF(COUNTA(A4:A303)>0,COUNTIF(G4:G303,"Sí")/COUNTA(A4:A303)*100,0)',AZUL_CLAR,AZUL_OSC),("Reseñas recibidas",'=COUNTIF(K4:K303,"Sí")',VERDE_CLAR,VERDE_OSC),("Clientes recurrentes",'=COUNTIF(L4:L303,"Sí")',VERDE_CLAR,VERDE_OSC),("Fuente: Facebook",'=COUNTIF(D4:D303,"Facebook")',GRIS_CLAR,GRIS_OSC),("Fuente: WhatsApp",'=COUNTIF(D4:D303,"WhatsApp")',GRIS_CLAR,GRIS_OSC),("Fuente: Recomendación",'=COUNTIF(D4:D303,"Recomendación")',GRIS_CLAR,GRIS_OSC),("Fuente: Saleads",'=COUNTIF(D4:D303,"Saleads")',GRIS_CLAR,GRIS_OSC)]
    for idx,(label,formula,bg_c,fg_c) in enumerate(crm,3):
        rr=idx; c1=ws.cell(row=rr,column=col_m,value=label); c1.fill=fill(bg_c); c1.font=font(bold=(formula is None),color=fg_c); c1.border=border_thin(); c1.alignment=align(h="left"); row_h(ws,rr,22)
        if formula:
            c2=ws.cell(row=rr,column=col_m+1,value=formula); c2.fill=fill(bg_c); c2.font=font(bold=True,color=fg_c); c2.border=border_thin(); c2.number_format=('0.0"%"' if "%" in label else '#,##0'); c2.alignment=align(h="right")
    return ws


# ── HISTORIAL DE MANTENIMIENTO (NUEVA HOJA) ───────────────────────────────────
def build_mantenimiento(wb):
    ws = wb.create_sheet("Mantenimiento")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    merge_header(ws,1,1,9,"MUDANZAS METEPEC — HISTORIAL DE MANTENIMIENTO",bg=ROJO_OSC,size=14)
    row_h(ws,1,35)
    merge_header(ws,2,1,9,"Registro completo por unidad · Conectado con Flujo de Efectivo",bg="C0392B",size=10)
    row_h(ws,2,22)

    headers=[("FOLIO MTO",11),("FECHA",12),("UNIDAD",16),("TIPO",22),
             ("KM EN MTO",12),("COSTO $",13),("PROVEEDOR",22),
             ("FOLIO GASTO",12),("OBSERVACIONES",35)]
    for i,(h,w) in enumerate(headers,1):
        header_cell(ws,3,i,h,bg=ROJO_OSC); col_w(ws,i,w)
    row_h(ws,3,30)

    for r in range(4,304):
        rb=BLANCO if r%2==0 else GRIS_CLAR

        # Folio automático
        c=ws.cell(row=r,column=1,value=f'=IF(B{r}<>"","MTO-"&TEXT(ROW()-3,"000"),"")')
        c.fill=fill(ROJO_CLAR); c.font=font(bold=True,color=ROJO_OSC)
        c.border=border_thin(); c.alignment=align(h="center")

        for col in [2]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.number_format="DD/MM/YYYY"; c.alignment=align(h="center")

        for col in [3,4,7]:
            c=ws.cell(row=r,column=col); c.fill=fill(rb); c.border=border_thin()
            c.alignment=align(h="center" if col!=7 else "left")

        c=ws.cell(row=r,column=5); c.fill=fill(rb); c.border=border_thin()
        c.number_format='#,##0" km"'; c.alignment=align(h="right")

        c=ws.cell(row=r,column=6); c.fill=fill(ROJO_CLAR); c.font=font(color=ROJO_OSC)
        c.border=border_thin(); c.number_format='"$"#,##0.00'; c.alignment=align(h="right")

        c=ws.cell(row=r,column=8); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="center")

        c=ws.cell(row=r,column=9); c.fill=fill(rb); c.border=border_thin()
        c.alignment=align(h="left",wrap=True)

        row_h(ws,r,18)

    add_dropdown(ws,'"Pick Up,1.5 Cerrada,2.5,3.5"',"C4:C303")
    add_dropdown(ws,'"Cambio aceite,Cambio llantas,Frenos,Afinación,Batería,Verificación,Suspensión,Transmisión,Motor,Carrocería,Eléctrico,Preventivo,Otro"',"D4:D303")

    # == PANEL RESUMEN POR UNIDAD ==
    col_s = 11
    ws.column_dimensions[get_column_letter(col_s)].width=26
    ws.column_dimensions[get_column_letter(col_s+1)].width=16

    merge_header(ws,3,col_s,col_s+1,"RESUMEN COSTOS MTO",bg=ROJO_OSC,size=10)

    mto_items=[
        ("TOTAL FLOTA",       '=SUM(F4:F303)',ROJO_OSC,BLANCO,True),
        ("Pick Up",           '=SUMIF(C4:C303,"Pick Up",F4:F303)',ROJO_CLAR,ROJO_OSC,False),
        ("1.5 Cerrada",       '=SUMIF(C4:C303,"1.5 Cerrada",F4:F303)',ROJO_CLAR,ROJO_OSC,False),
        ("2.5",               '=SUMIF(C4:C303,"2.5",F4:F303)',ROJO_CLAR,ROJO_OSC,False),
        ("3.5",               '=SUMIF(C4:C303,"3.5",F4:F303)',ROJO_CLAR,ROJO_OSC,False),
        ("",                  None,BLANCO,NEGRO,False),
        ("Cambio aceite",     '=SUMIF(D4:D303,"Cambio aceite",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("Cambio llantas",    '=SUMIF(D4:D303,"Cambio llantas",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("Frenos",            '=SUMIF(D4:D303,"Frenos",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("Afinación",         '=SUMIF(D4:D303,"Afinación",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("Motor",             '=SUMIF(D4:D303,"Motor",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("Preventivo",        '=SUMIF(D4:D303,"Preventivo",F4:F303)',GRIS_CLAR,GRIS_OSC,False),
        ("",                  None,BLANCO,NEGRO,False),
        ("Nº registros",      '=COUNTA(B4:B303)',AZUL_CLAR,AZUL_OSC,True),
        ("Últ. registro",     '=IFERROR(MAX(B4:B303),"—")',AZUL_CLAR,AZUL_OSC,False),
        ("Prom. x evento $",  '=IFERROR(SUM(F4:F303)/COUNTA(B4:B303),0)',AZUL_CLAR,AZUL_OSC,False),
    ]

    for idx,(label,formula,bg_c,fg_c,bold_) in enumerate(mto_items,4):
        rr=idx
        c1=ws.cell(row=rr,column=col_s,value=label)
        c1.fill=fill(bg_c); c1.font=font(bold=bold_,color=fg_c)
        c1.border=border_thin(); c1.alignment=align(h="left")
        row_h(ws,rr,22)
        if formula:
            c2=ws.cell(row=rr,column=col_s+1,value=formula)
            c2.fill=fill(bg_c); c2.font=font(bold=bold_,color=fg_c)
            c2.border=border_thin()
            c2.number_format='"$"#,##0.00' if label!="Nº registros" else '#,##0'
            if label=="Últ. registro": c2.number_format="DD/MM/YYYY"
            if label=="Nº registros": c2.number_format='#,##0'
            c2.alignment=align(h="right")

    return ws


# ── DASHBOARD (v1 íntegra + bloque adicional) ─────────────────────────────────
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    merge_header(ws,1,1,12,"MUDANZAS METEPEC — DASHBOARD OPERATIVO",bg=AZUL_OSC,size=16)
    row_h(ws,1,45)
    ws.merge_cells("A2:L2")
    c2=ws.cell(row=2,column=1,value='="Actualizado: "&TEXT(TODAY(),"DD/MM/YYYY")')
    c2.fill=fill(AZUL_MED); c2.font=font(bold=True,size=11,color=BLANCO); c2.alignment=align(h="right")
    row_h(ws,2,22)
    col_w(ws,1,2)
    for c in range(2,13): col_w(ws,c,16)

    def kpi_block(ws,row,col,title,formula,bg_title,bg_value,fg_v,num_fmt=None,rh=40):
        ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1)
        t=ws.cell(row=row,column=col,value=title); t.fill=fill(bg_title); t.font=font(bold=True,size=9,color=BLANCO); t.alignment=align(h="center"); t.border=border_thin()
        ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
        v=ws.cell(row=row+1,column=col,value=formula); v.fill=fill(bg_value); v.font=font(bold=True,size=18,color=fg_v); v.alignment=align(h="center"); v.border=border_thin()
        if num_fmt: v.number_format=num_fmt
        row_h(ws,row+1,rh); return v

    # ── BLOQUE v1 intacto ──
    r=4; row_h(ws,r,22); row_h(ws,r+1,45)
    ws.merge_cells(f"B{r}:M{r}")
    sec=ws.cell(row=r,column=2,value="SERVICIOS"); sec.fill=fill(GRIS_OSC); sec.font=font(bold=True,color=BLANCO,size=11); sec.alignment=align(h="center"); sec.border=border_thin()
    r+=1
    kpi_sv=[("Servicios HOY",'=COUNTIF(Agenda!C:C,TODAY())',AZUL_OSC,AZUL_CLAR,AZUL_OSC),("Esta SEMANA",'=COUNTIFS(Agenda!C:C,">="&TODAY()-WEEKDAY(TODAY(),2)+1,Agenda!C:C,"<="&TODAY()-WEEKDAY(TODAY(),2)+7)',AZUL_MED,AZUL_CLAR,AZUL_OSC),("Este MES",'=COUNTIFS(Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))',AZUL_OSC,AZUL_CLAR,AZUL_OSC),("PENDIENTES",'=COUNTIF(Agenda!Q:Q,"Pendiente")',AMBAR,AMBAR_CLAR,AMBAR),("CONFIRMADOS",'=COUNTIF(Agenda!Q:Q,"Confirmado")',VERDE_MED,VERDE_CLAR,VERDE_OSC),("TERMINADOS mes",'=COUNTIFS(Agenda!Q:Q,"Terminado",Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))',VERDE_OSC,VERDE_CLAR,VERDE_OSC)]
    cp=2
    for title,formula,bg_t,bg_v,fg_v in kpi_sv: kpi_block(ws,r,cp,title,formula,bg_t,bg_v,fg_v); cp+=2
    r+=2; row_h(ws,r,12); r+=1
    ws.merge_cells(f"B{r}:M{r}")
    sec2=ws.cell(row=r,column=2,value="FINANZAS DEL MES"); sec2.fill=fill(VERDE_OSC); sec2.font=font(bold=True,color=BLANCO,size=11); sec2.alignment=align(h="center"); sec2.border=border_thin()
    row_h(ws,r,22); r+=1; row_h(ws,r+1,45)
    kpi_fin=[("VENTAS MES",'=IFERROR(SUMIFS(Agenda!N:N,Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0)),0)',VERDE_OSC,VERDE_CLAR,VERDE_OSC,'"$"#,##0'),("COSTOS MES",'=IFERROR(SUMIFS(Costos!P:P,Costos!P:P,">"&0),0)',ROJO_OSC,ROJO_CLAR,ROJO_OSC,'"$"#,##0'),("UTILIDAD MES",'=IFERROR(SUMIFS(Agenda!N:N,Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))-SUMIF(Costos!P:P,">"&0,Costos!P:P),0)',VERDE_OSC,VERDE_CLAR,VERDE_OSC,'"$"#,##0'),("GASTOS FIJOS",'=IFERROR(\'Gastos Fijos\'!B11,0)',ROJO_OSC,ROJO_CLAR,ROJO_OSC,'"$"#,##0'),("FLUJO NETO",'=IFERROR(SUM(Flujo!C4:C504)-SUM(Flujo!D4:D504),0)',AZUL_OSC,AZUL_CLAR,AZUL_OSC,'"$"#,##0'),("MARGEN PROM %",'=IFERROR(AVERAGE(Costos!S4:S203),0)',AZUL_MED,AZUL_CLAR,AZUL_OSC,'0.0"%"')]
    cp=2
    for title,formula,bg_t,bg_v,fg_v,nfmt in kpi_fin: kpi_block(ws,r,cp,title,formula,bg_t,bg_v,fg_v,num_fmt=nfmt); cp+=2
    r+=2; row_h(ws,r,12); r+=1
    ws.merge_cells(f"B{r}:M{r}")
    sec3=ws.cell(row=r,column=2,value="VENTAS POR UNIDAD — MES ACTUAL"); sec3.fill=fill(AZUL_OSC); sec3.font=font(bold=True,color=BLANCO,size=11); sec3.alignment=align(h="center"); sec3.border=border_thin()
    row_h(ws,r,22); r+=1
    unit_h=["UNIDAD","SERVICIOS","VENTAS $","META MES $","CUMP %","FALTA $","UTILIDAD $"]
    for i,h in enumerate(unit_h,2): header_cell(ws,r,i,h,bg=AZUL_MED); row_h(ws,r,26)
    r+=1
    units_dash=[("Pick Up",24250,AMBAR_CLAR),("1.5 Cerrada",28450,VERDE_CLAR),("2.5",31250,AZUL_CLAR),("3.5",15250,GRIS_CLAR)]
    for uname,meta,bg_u in units_dash:
        uq=f'"{uname}"'; rb=bg_u
        ws.cell(row=r,column=2,value=uname).fill=fill(rb); ws.cell(row=r,column=2).font=font(bold=True,color=GRIS_OSC); ws.cell(row=r,column=2).border=border_thin(); ws.cell(row=r,column=2).alignment=align(h="center")
        sv=ws.cell(row=r,column=3,value=f'=COUNTIFS(Agenda!K:K,{uq},Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))'); sv.fill=fill(rb); sv.border=border_thin(); sv.alignment=align(h="center"); sv.font=font(bold=True)
        vt=ws.cell(row=r,column=4,value=f'=SUMIFS(Agenda!N:N,Agenda!K:K,{uq},Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),Agenda!C:C,"<="&EOMONTH(TODAY(),0))'); vt.fill=fill(rb); vt.border=border_thin(); vt.number_format='"$"#,##0.00'; vt.alignment=align(h="right"); vt.font=font(bold=True,color=VERDE_OSC)
        mt=ws.cell(row=r,column=5,value=meta); mt.fill=fill(rb); mt.border=border_thin(); mt.number_format='"$"#,##0.00'; mt.alignment=align(h="right"); mt.font=font(color=AMBAR)
        cp2=ws.cell(row=r,column=6,value=f'=IF(E{r}>0,D{r}/E{r}*100,0)'); cp2.fill=fill(rb); cp2.border=border_thin(); cp2.number_format='0.0"%"'; cp2.alignment=align(h="right"); cp2.font=font(bold=True)
        fl=ws.cell(row=r,column=7,value=f'=MAX(0,E{r}-D{r})'); fl.fill=fill(rb); fl.border=border_thin(); fl.number_format='"$"#,##0.00'; fl.alignment=align(h="right"); fl.font=font(color=ROJO_OSC)
        ut=ws.cell(row=r,column=8,value=f'=SUMIF(Costos!C:C,{uq},Costos!R:R)'); ut.fill=fill(rb); ut.border=border_thin(); ut.number_format='"$"#,##0.00'; ut.alignment=align(h="right"); ut.font=font(bold=True,color=VERDE_OSC)
        row_h(ws,r,22); r+=1
    r+=1
    ws.merge_cells(f"B{r}:M{r}")
    sec4=ws.cell(row=r,column=2,value="CRM — CONVERSIÓN DE CLIENTES"); sec4.fill=fill(GRIS_OSC); sec4.font=font(bold=True,color=BLANCO,size=11); sec4.alignment=align(h="center"); sec4.border=border_thin()
    row_h(ws,r,22); r+=1; row_h(ws,r+1,45)
    kpi_crm=[("CONTACTOS",'=COUNTA(Clientes!A4:A303)',AZUL_OSC,AZUL_CLAR,AZUL_OSC,None),("COTIZACIONES",'=COUNTA(Clientes!F4:F303)',AZUL_MED,AZUL_CLAR,AZUL_OSC,None),("CERRADOS",'=COUNTIF(Clientes!G4:G303,"Sí")',VERDE_MED,VERDE_CLAR,VERDE_OSC,None),("TASA CIERRE",'=IF(COUNTA(Clientes!A4:A303)>0,COUNTIF(Clientes!G4:G303,"Sí")/COUNTA(Clientes!A4:A303)*100,0)',VERDE_OSC,VERDE_CLAR,VERDE_OSC,'0.0"%"'),("RESEÑAS",'=COUNTIF(Clientes!K4:K303,"Sí")',AMBAR,AMBAR_CLAR,AMBAR,None),("RECURRENTES",'=COUNTIF(Clientes!L4:L303,"Sí")',VERDE_OSC,VERDE_CLAR,VERDE_OSC,None)]
    cp=2
    for title,formula,bg_t,bg_v,fg_v,nfmt in kpi_crm: kpi_block(ws,r,cp,title,formula,bg_t,bg_v,fg_v,num_fmt=nfmt); cp+=2
    r+=2; row_h(ws,r,12); r+=1
    ws.merge_cells(f"B{r}:G{r}")
    sec5=ws.cell(row=r,column=2,value="COBERTURA GASTOS FIJOS"); sec5.fill=fill(ROJO_OSC); sec5.font=font(bold=True,color=BLANCO,size=11); sec5.alignment=align(h="center"); sec5.border=border_thin()
    ws.merge_cells(f"H{r}:M{r}")
    sec6=ws.cell(row=r,column=8,value="MEJOR UNIDAD DEL MES"); sec6.fill=fill(VERDE_OSC); sec6.font=font(bold=True,color=BLANCO,size=11); sec6.alignment=align(h="center"); sec6.border=border_thin()
    row_h(ws,r,22); r+=1
    ws.merge_cells(f"B{r}:G{r}")
    cob=ws.cell(row=r,column=2,value=f'="Ventas: $"&TEXT(IFERROR(SUM(Agenda!N:N),0),"#,##0")&" / Meta: $"&TEXT(IFERROR(\'Gastos Fijos\'!B11,0),"#,##0")&" ("&TEXT(IFERROR(SUM(Agenda!N:N)/\'Gastos Fijos\'!B11*100,0),"0.0")&"%)"')
    cob.fill=fill(ROJO_CLAR); cob.font=font(bold=True,size=12,color=ROJO_OSC); cob.alignment=align(h="center"); cob.border=border_thin()
    row_h(ws,r,38)
    ws.merge_cells(f"H{r}:M{r}")
    mejor=ws.cell(row=r,column=8,
        value='=IFERROR(INDEX({"Pick Up","1.5 Cerrada","2.5","3.5"},MATCH(MAX(SUMIF(Agenda!K:K,"Pick Up",Agenda!N:N),SUMIF(Agenda!K:K,"1.5 Cerrada",Agenda!N:N),SUMIF(Agenda!K:K,"2.5",Agenda!N:N),SUMIF(Agenda!K:K,"3.5",Agenda!N:N)),{SUMIF(Agenda!K:K,"Pick Up",Agenda!N:N),SUMIF(Agenda!K:K,"1.5 Cerrada",Agenda!N:N),SUMIF(Agenda!K:K,"2.5",Agenda!N:N),SUMIF(Agenda!K:K,"3.5",Agenda!N:N)},0)),"—")')
    mejor.fill=fill(VERDE_CLAR); mejor.font=font(bold=True,size=16,color=VERDE_OSC); mejor.alignment=align(h="center"); mejor.border=border_thin()
    r+=1

    # ── BLOQUE NUEVO: indicadores adicionales ──
    r+=1; row_h(ws,r,12); r+=1
    ws.merge_cells(f"B{r}:M{r}")
    sec_new=ws.cell(row=r,column=2,value="OPERACIÓN Y RENTABILIDAD")
    sec_new.fill=fill(AZUL_OSC); sec_new.font=font(bold=True,color=BLANCO,size=11)
    sec_new.alignment=align(h="center"); sec_new.border=border_thin()
    row_h(ws,r,22); r+=1; row_h(ws,r+1,45)

    kpi_op=[
        ("CANCELADOS mes",'=COUNTIFS(Agenda!Q:Q,"Cancelado",Agenda!C:C,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))',ROJO_OSC,ROJO_CLAR,ROJO_OSC,None),
        ("EN PROCESO",'=COUNTIF(Agenda!Q:Q,"En proceso")',AMBAR,AMBAR_CLAR,AMBAR,None),
        ("COSTO PROM./SV",'=IFERROR(SUM(Costos!P4:P203)/COUNTA(Costos!A4:A203),0)',ROJO_MED if False else ROJO_OSC,ROJO_CLAR,ROJO_OSC,'"$"#,##0'),
        ("INGRESO PROM./SV",'=IFERROR(SUM(Agenda!N4:N203)/COUNTA(Agenda!E4:E203),0)',VERDE_OSC,VERDE_CLAR,VERDE_OSC,'"$"#,##0'),
        ("COSTO MTO TOTAL",'=IFERROR(SUM(Mantenimiento!F4:F303),0)',ROJO_OSC,ROJO_CLAR,ROJO_OSC,'"$"#,##0'),
        ("UNIDAD MÁS OCUP.",'=IFERROR(INDEX({"Pick Up","1.5 Cerrada","2.5","3.5"},MATCH(MAX(COUNTIFS(Agenda!K:K,"Pick Up",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"1.5 Cerrada",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"2.5",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"3.5",Agenda!Q:Q,"Terminado")),{COUNTIFS(Agenda!K:K,"Pick Up",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"1.5 Cerrada",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"2.5",Agenda!Q:Q,"Terminado"),COUNTIFS(Agenda!K:K,"3.5",Agenda!Q:Q,"Terminado")},0)),"—")',VERDE_MED,VERDE_CLAR,VERDE_OSC,None),
    ]

    cp=2
    for title,formula,bg_t,bg_v,fg_v,nfmt in kpi_op:
        kpi_block(ws,r,cp,title,formula,bg_t,bg_v,fg_v,num_fmt=nfmt); cp+=2

    r+=2; row_h(ws,r,12); r+=1

    # Utilidad por unidad - mini tabla
    ws.merge_cells(f"B{r}:M{r}")
    sec_ut=ws.cell(row=r,column=2,value="UTILIDAD Y RENTABILIDAD POR UNIDAD (ACUMULADO)")
    sec_ut.fill=fill(VERDE_OSC); sec_ut.font=font(bold=True,color=BLANCO,size=11)
    sec_ut.alignment=align(h="center"); sec_ut.border=border_thin()
    row_h(ws,r,22); r+=1

    ut_headers=["UNIDAD","VENTAS $","COSTOS $","UTILIDAD $","MARGEN %","COSTO MTO $","UTIL.NETA $"]
    for i,h in enumerate(ut_headers,2): header_cell(ws,r,i,h,bg=VERDE_MED); row_h(ws,r,26)
    r+=1

    for uname,bg_u in [("Pick Up",AMBAR_CLAR),("1.5 Cerrada",VERDE_CLAR),("2.5",AZUL_CLAR),("3.5",GRIS_CLAR)]:
        uq=f'"{uname}"'
        ws.cell(row=r,column=2,value=uname).fill=fill(bg_u); ws.cell(row=r,column=2).font=font(bold=True,color=GRIS_OSC); ws.cell(row=r,column=2).border=border_thin(); ws.cell(row=r,column=2).alignment=align(h="center")
        vt=ws.cell(row=r,column=3,value=f'=SUMIF(Agenda!K:K,{uq},Agenda!N:N)'); vt.fill=fill(bg_u); vt.border=border_thin(); vt.number_format='"$"#,##0.00'; vt.alignment=align(h="right"); vt.font=font(bold=True,color=VERDE_OSC)
        ct=ws.cell(row=r,column=4,value=f'=SUMIF(Costos!C:C,{uq},Costos!P:P)'); ct.fill=fill(bg_u); ct.border=border_thin(); ct.number_format='"$"#,##0.00'; ct.alignment=align(h="right"); ct.font=font(color=ROJO_OSC)
        ut=ws.cell(row=r,column=5,value=f'=C{r}-D{r}'); ut.fill=fill(VERDE_CLAR); ut.border=border_thin(); ut.number_format='"$"#,##0.00'; ut.alignment=align(h="right"); ut.font=font(bold=True,color=VERDE_OSC)
        mg=ws.cell(row=r,column=6,value=f'=IF(C{r}>0,E{r}/C{r}*100,0)'); mg.fill=fill(bg_u); mg.border=border_thin(); mg.number_format='0.0"%"'; mg.alignment=align(h="right"); mg.font=font(bold=True)
        mto=ws.cell(row=r,column=7,value=f'=SUMIF(Mantenimiento!C:C,{uq},Mantenimiento!F:F)'); mto.fill=fill(ROJO_CLAR); mto.border=border_thin(); mto.number_format='"$"#,##0.00'; mto.alignment=align(h="right"); mto.font=font(color=ROJO_OSC)
        un=ws.cell(row=r,column=8,value=f'=E{r}-G{r}'); un.fill=fill(VERDE_CLAR); un.border=border_thin(); un.number_format='"$"#,##0.00'; un.alignment=align(h="right"); un.font=font(bold=True,color=VERDE_OSC)
        row_h(ws,r,22); r+=1

    return ws


# ── INSTRUCCIONES (v1 + nuevas secciones) ────────────────────────────────────
def build_instrucciones(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_view.showGridLines = False
    col_w(ws,1,4); col_w(ws,2,60); col_w(ws,3,40)
    merge_header(ws,1,1,3,"MUDANZAS METEPEC — GUÍA DE USO DEL SISTEMA",bg=AZUL_OSC,size=14)
    row_h(ws,1,40)
    instrucciones=[
        ("",BLANCO),
        ("FLUJO DE TRABAJO DIARIO:",AZUL_OSC),
        ("1. AGENDA → Registra cada servicio desde WhatsApp. El folio se genera solo.",BLANCO),
        ("2. AGENDA → Llena: cliente, fecha, hora, tipo, origen, destino, unidad, precio cotizado, anticipo.",GRIS_CLAR),
        ("3. AGENDA → Cambia el Estado conforme avanza: Pendiente → Confirmado → En proceso → Terminado.",BLANCO),
        ("4. AGENDA → Llena las columnas nuevas: Prioridad, Fuente, Tipo cliente, Horas, Fotos/Videos.",GRIS_CLAR),
        ("5. COSTOS → Copia el Folio de la Agenda. El cliente y la unidad se llenan solos.",BLANCO),
        ("6. COSTOS → Llena kilómetros reales, chofer, ayudantes, maniobras, casetas, comidas, etc.",GRIS_CLAR),
        ("7. FLUJO → Registra cada entrada de dinero (anticipo, saldo) y cada gasto del día.",BLANCO),
        ("8. MANTENIMIENTO → Cada vez que hagas mantenimiento a una unidad, regístralo aquí.",GRIS_CLAR),
        ("9. UNIDADES → Actualiza el Kilometraje actual de cada unidad al menos una vez por semana.",BLANCO),
        ("10. CLIENTES → Registra cada contacto nuevo. Llena si cerró, fuente y seguimiento.",GRIS_CLAR),
        ("11. DASHBOARD → Revisa cada mañana el estado del negocio en 30 segundos.",BLANCO),
        ("",BLANCO),
        ("COLUMNAS NUEVAS EN AGENDA (cols 23-37):",AZUL_OSC),
        ("Prioridad: Alta (urgente) / Media / Baja — colorea la fila automáticamente.",BLANCO),
        ("Horas Salida/Llegada/Término — el Tiempo Real se calcula solo.",GRIS_CLAR),
        ("Fotos, Videos, Reseña, Permiso publicar — control de contenido digital.",BLANCO),
        ("Fuente del cliente — de dónde llegó: Facebook, TikTok, Google, Recomendación…",GRIS_CLAR),
        ("Tipo de cliente: Particular, Empresa, Recurrente, VIP",BLANCO),
        ("",BLANCO),
        ("EXPEDIENTE DE UNIDADES (hoja Unidades):",AZUL_OSC),
        ("Llena los datos técnicos de cada camioneta UNA SOLA VEZ.",BLANCO),
        ("El % de vida útil y KM restantes se calculan solos.",GRIS_CLAR),
        ("Actualiza solo el Kilometraje actual cada semana.",BLANCO),
        ("Las fechas de seguro, verificación y tenencia generan alertas automáticas.",GRIS_CLAR),
        ("",BLANCO),
        ("MANTENIMIENTO:",ROJO_OSC),
        ("Registra CADA mantenimiento: fecha, tipo, costo, km y proveedor.",ROJO_CLAR),
        ("El resumen por unidad y por tipo se calcula automáticamente.",ROJO_CLAR),
        ("Este costo aparece en el Dashboard como Utilidad Neta (Utilidad - Costo MTO).",ROJO_CLAR),
        ("",BLANCO),
        ("METAS MENSUALES:",AZUL_OSC),
        ("Pick Up: $24,250 | 1.5 Cerrada: $28,450 | 2.5: $31,250 | 3.5: $15,250",AMBAR_CLAR),
        ("Total flota: $99,200 (igual que gastos fijos — ahí empieza la utilidad)",AMBAR_CLAR),
        ("",BLANCO),
        ("METAS SEMANALES:",AZUL_OSC),
        ("Pick Up: $5,600 | 1.5 Cerrada: $6,570 | 2.5: $7,220 | 3.5: $3,520",VERDE_CLAR),
        ("",BLANCO),
        ("COSTO POR KM (automático en Costos):",AZUL_OSC),
        ("Pick Up: $7/km  |  1.5 Cerrada: $9/km  |  2.5: $10/km  |  3.5: $12/km",GRIS_CLAR),
        ("",BLANCO),
        ("TIPS IMPORTANTES:",ROJO_OSC),
        ("- El folio conecta TODO: Agenda → Costos → Flujo → Mantenimiento.",ROJO_CLAR),
        ("- Usa siempre los dropdowns — los cálculos dependen del texto exacto.",ROJO_CLAR),
        ("- Agrega el anticipo el mismo día que lo recibes en Flujo.",ROJO_CLAR),
        ("- El Dashboard muestra Utilidad Neta = Utilidad operativa − Costos de mantenimiento.",ROJO_CLAR),
        ("",BLANCO),
        ("GASTOS FIJOS MENSUALES TOTALES: $99,200",VERDE_OSC),
        ("Todo lo que supere $99,200 en ventas es utilidad real del negocio.",VERDE_CLAR),
    ]
    for idx,(text,bg_c) in enumerate(instrucciones,2):
        rr=idx
        ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=3)
        c=ws.cell(row=rr,column=2,value=text)
        is_h=bg_c in [AZUL_OSC,VERDE_OSC,ROJO_OSC]
        c.fill=fill(bg_c if bg_c!=BLANCO else BLANCO)
        c.font=font(bold=is_h,size=11 if is_h else 10,color=BLANCO if is_h else (GRIS_OSC if bg_c==GRIS_CLAR else NEGRO))
        c.alignment=align(h="left",v="center"); c.border=border_thin()
        row_h(ws,rr,22 if not is_h else 28)
        ws.cell(row=rr,column=1).fill=fill(AZUL_OSC if is_h else AZUL_MED)
        ws.cell(row=rr,column=1).border=border_thin()
    return ws


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)

    print("Construyendo hojas...")
    build_dashboard(wb)
    build_agenda(wb)
    build_costos(wb)
    build_unidades(wb)
    build_flujo(wb)
    build_gastos_fijos(wb)
    build_clientes(wb)
    build_mantenimiento(wb)
    build_instrucciones(wb)

    wb.active = wb["Dashboard"]
    out = "MudanzasMetepec_Sistema_v2.xlsx"
    wb.save(out)
    print(f"\n✅  Generado: {out}")
    print("    Backup v1 en: MudanzasMetepec_Sistema_BACKUP_v1.xlsx")

if __name__ == "__main__":
    main()
